import base64
import json
import mimetypes
import os
import re
from io import BytesIO

from flask import Flask, request, jsonify, send_file, Response

app = Flask(__name__)


# =============================================================
#  Utilidades
# =============================================================
def num(v, d=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def inum(v, d=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return d


def money(n):
    """Formato pesos dominicanos: RD$ 24,000"""
    s = f"{int(round(n or 0)):,}"
    return f"RD$ {s}"


def latin(s):
    """Deja el texto seguro para las fuentes base del PDF."""
    return str(s or "").encode("latin-1", "replace").decode("latin-1")


DRINK_KEYWORDS = [
    "gaseosa", "cerveza", "cola", "malta", "jugo", "agua", "refresco",
    "limonada", "te", "café", "cafe", "poker", "aguila", "águila",
    "soda", "bebida", "vino", "whisky", "aguardiente", "ron", "michelada",
    "presidente", "heineken", "corona", "modelo", "brahma", "stella",
    "paulaner", "pellegrino", "sprite", "cocacola", "coca cola",
]


def guess_kind(desc):
    d = (desc or "").lower()
    return "individual" if any(k in d for k in DRINK_KEYWORDS) else "shared"


# =============================================================
#  Cálculo (fuente única de verdad)
# =============================================================
def compute(state):
    event = state.get("event", {}) or {}
    pv = num(event.get("poldinaValue"), 1000) or 1000
    people = state.get("people", []) or []
    fines = state.get("fines", []) or []
    items = state.get("items", []) or []

    fines_by_person = {}
    for f in fines:
        pid = f.get("personId")
        fines_by_person[pid] = fines_by_person.get(pid, 0) + 1

    total_pold = sum(fines_by_person.values())
    pool = total_pold * pv

    shared_items = [i for i in items if i.get("kind") != "individual"]
    indiv_items = [i for i in items if i.get("kind") == "individual"]
    indiv_assigned = [i for i in indiv_items if i.get("personId")]
    indiv_unassigned = [i for i in indiv_items if not i.get("personId")]

    shared_cost = sum(num(i.get("qty")) * num(i.get("unitPrice")) for i in shared_items)
    excess = max(0.0, shared_cost - pool)
    surplus = max(0.0, pool - shared_cost)
    covered = min(shared_cost, pool)

    excess_per = (excess / len(people)) if people else 0.0

    rows = []
    for p in people:
        pold = fines_by_person.get(p.get("id"), 0)
        fine = pold * pv
        ex = sum(
            num(i.get("qty")) * num(i.get("unitPrice"))
            for i in indiv_assigned if i.get("personId") == p.get("id")
        )
        eshare = excess_per if excess > 0 else 0.0
        rows.append({
            "id": p.get("id"),
            "name": p.get("name") or "(sin nombre)",
            "poldinas": pold,
            "fine": fine,
            "excessShare": eshare,
            "extras": ex,
            "total": fine + eshare + ex,
            "fined": pold > 0,
        })

    extras_total = sum(num(i.get("qty")) * num(i.get("unitPrice")) for i in indiv_assigned)
    unassigned_total = sum(num(i.get("qty")) * num(i.get("unitPrice")) for i in indiv_unassigned)
    grand = sum(r["total"] for r in rows)

    return {
        "pv": pv,
        "totalPoldinas": total_pold,
        "pool": pool,
        "sharedCost": shared_cost,
        "covered": covered,
        "excess": excess,
        "surplus": surplus,
        "extrasTotal": extras_total,
        "unassignedTotal": unassigned_total,
        "unassignedCount": len(indiv_unassigned),
        "grand": grand,
        "excessPer": excess_per,
        "rows": rows,
    }


# =============================================================
#  Lectura de factura: pegar texto (sin IA, siempre disponible)
# =============================================================
PATTERNS = [
    re.compile(r"^\s*(?P<qty>\d+)\s*[xX]\s*(?P<desc>.+?)\s+\$?\s*(?P<price>[\d.,]+)\s*$"),   # 2x Pizza mediana 24000
    re.compile(r"^\s*(?P<desc>.+?)\s+[xX]\s*(?P<qty>\d+)\s+\$?\s*(?P<price>[\d.,]+)\s*$"),   # Pizza mediana x2 24000
    re.compile(r"^\s*(?P<qty>\d+)\s+(?P<desc>.+?)\s+\$?\s*(?P<price>[\d.,]+)\s*$"),          # 2 Pizza mediana 24000
    re.compile(r"^\s*(?P<desc>.+?)\s+\$?\s*(?P<price>[\d.,]+)\s*$"),                          # Pizza mediana 24000
]


SKIP_KEYWORDS = [
    "total", "subtotal", "base imponible", "itbis", "descuento", "cambio",
    "entregado", "propina", "servicio", "impuesto", "rnc", "ncf", "factura",
    "mesa", "mesero", "cajero", "cliente", "salonero", "hora", "fecha",
    "forma de pago", "ley ",
]


def parse_invoice_text(text):
    items = []
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        low = line.lower()
        if any(k in low for k in SKIP_KEYWORDS):
            continue
        m = None
        for pat in PATTERNS:
            m = pat.match(line)
            if m:
                break
        if not m:
            continue
        gd = m.groupdict()
        qty = int(gd.get("qty") or 1)
        price_raw = gd["price"].replace(",", "")
        try:
            price = float(price_raw)
        except ValueError:
            continue
        desc = gd["desc"].strip(" -:\t")
        if not desc or price <= 0:
            continue
        unit_price = price / qty if qty else price
        items.append({
            "desc": desc,
            "qty": qty,
            "unitPrice": round(unit_price),
            "kind": guess_kind(desc),
        })
    return items


DATE_RE = re.compile(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})")


def _to_number(s):
    return float(s.replace(",", ""))


def detect_totals_and_meta(text):
    meta = {}
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    if lines:
        meta["place"] = lines[0][:80]
    for line in lines:
        low = line.lower()
        num_m = re.search(r"([\d.,]+)\s*$", line)
        if not num_m:
            continue
        try:
            value = _to_number(num_m.group(1))
        except ValueError:
            continue
        if ("subtotal" in low or "base imponible" in low) and "subtotal" not in meta:
            meta["subtotal"] = value
        elif low.startswith("total") and "total" not in meta:
            meta["total"] = value
    dm = DATE_RE.search(text or "")
    if dm:
        d, mo, y = dm.groups()
        y = ("20" + y) if len(y) == 2 else y
        try:
            meta["date"] = f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            pass
    return meta


# =============================================================
#  Lectura de PDF con texto real — gratis, sin IA
#  (solo funciona si el PDF tiene texto de verdad adentro, no si es
#  una foto/escaneo guardado como PDF)
# =============================================================
def extract_pdf_text(file_bytes):
    import pdfplumber

    parts = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t:
                parts.append(t)
    return "\n".join(parts).strip()


# =============================================================
#  Lectura de fotos con OCR local — gratis, sin IA, sin API
#  (usa RapidOCR: se instala solo con pip, no necesita ningun
#  programa de sistema como Tesseract, asi que es compatible con
#  entornos serverless como Vercel)
# =============================================================
class OCRNotAvailable(Exception):
    pass


_OCR_ENGINE = None


def _get_ocr_engine():
    global _OCR_ENGINE
    if _OCR_ENGINE is None:
        from rapidocr_onnxruntime import RapidOCR
        _OCR_ENGINE = RapidOCR()
    return _OCR_ENGINE


def ocr_image_text(file_bytes):
    """Lee el texto de una foto usando un motor de OCR local (sin llamadas externas)."""
    from PIL import Image, ImageOps
    import numpy as np

    try:
        img = Image.open(BytesIO(file_bytes))
        img = ImageOps.exif_transpose(img)  # corrige la rotacion que guardan los celulares
        img = img.convert("RGB")
    except Exception as e:
        raise ValueError(f"No se pudo abrir la imagen: {e}")

    w, h = img.size
    if max(w, h) < 1500:
        scale = 1500 / max(w, h)
        img = img.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.LANCZOS)

    try:
        engine = _get_ocr_engine()
        result, _ = engine(np.array(img))
    except Exception:
        raise OCRNotAvailable()

    if not result:
        return ""
    lines = [line[1] for line in result if len(line) > 1]
    return "\n".join(lines).strip()


def ocr_pdf_as_image(file_bytes):
    """Convierte cada pagina de un PDF sin texto real (un escaneo) a imagen y le aplica OCR."""
    import fitz  # PyMuPDF

    texts = []
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    try:
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            t = ocr_image_text(pix.tobytes("png"))
            if t:
                texts.append(t)
    finally:
        doc.close()
    return "\n".join(texts).strip()


# =============================================================
#  Lectura de factura con IA (foto o PDF) — requiere ANTHROPIC_API_KEY
# =============================================================
class MissingAPIKey(Exception):
    pass


def _extract_json_object(text):
    text = (text or "").strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    start, end = text.find("{"), text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError("La respuesta no contenía JSON.")
    return json.loads(text[start:end + 1])


def extract_invoice_data(file_bytes, mime_type):
    import requests

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise MissingAPIKey()

    model = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-5")
    b64 = base64.b64encode(file_bytes).decode()
    block = (
        {"type": "document", "source": {"type": "base64", "media_type": mime_type, "data": b64}}
        if mime_type == "application/pdf" else
        {"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": b64}}
    )

    prompt = (
        "Analiza esta factura o recibo de consumo. Responde SOLO con un JSON valido, sin texto adicional ni "
        "comentarios, con esta forma exacta: "
        '{"place":"nombre del establecimiento","date":"YYYY-MM-DD","subtotal":0,"total":0,'
        '"items":[{"desc":"nombre del item","qty":1,"unitPrice":0,"kind":"shared"}]}. '
        "place es el nombre del negocio en el encabezado del recibo. "
        "date es la fecha del consumo, convertida a formato YYYY-MM-DD. "
        "subtotal es el monto antes de impuestos (base imponible) y total es el monto final a pagar, "
        "incluyendo impuestos y cargos. Si no encuentras alguno de estos datos, usa null. "
        'Usa kind="shared" para comida para compartir (pizzas, platos grandes, entradas para la mesa) y '
        'kind="individual" para bebidas o platos individuales (cualquier bebida, cerveza, vino, refresco, '
        "agua, jugo, coctel, o plato que normalmente pide y come una sola persona). "
        "qty es la cantidad (numero entero) y unitPrice el precio unitario en pesos, sin simbolos ni "
        "separadores de miles. Si el recibo muestra el precio total de la linea en vez del unitario, "
        "calcula el unitario dividiendo el total entre la cantidad. "
        "No incluyas como item las lineas de total, subtotal, impuestos, descuento, propina o cambio."
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": model,
            "max_tokens": 1800,
            "messages": [{"role": "user", "content": [block, {"type": "text", "text": prompt}]}],
        },
        timeout=45,
    )
    resp.raise_for_status()
    data = resp.json()
    text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
    return _extract_json_object(text)


def guess_mime(file_storage):
    mime = file_storage.mimetype
    if mime in ("image/jpeg", "image/png", "image/webp", "image/gif", "application/pdf"):
        return mime
    guessed, _ = mimetypes.guess_type(file_storage.filename or "")
    return guessed or mime


# =============================================================
#  PDF
# =============================================================
def build_pdf(state):
    from fpdf import FPDF

    c = compute(state)
    event = state.get("event", {}) or {}
    name_of = {p.get("id"): (p.get("name") or "(sin nombre)") for p in state.get("people", [])}

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(23, 53, 44)
    pdf.cell(0, 10, latin("Poldinas - reparto de la cuenta"), new_x="LMARGIN", new_y="NEXT")

    meta = "   |   ".join([x for x in [event.get("name"), event.get("place"), event.get("date")] if x])
    if meta:
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(84, 101, 92)
        pdf.cell(0, 7, latin(meta), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.set_text_color(23, 53, 44)
    pairs = [
        ("Bolsa de multas", f"{c['totalPoldinas']} poldinas = {money(c['pool'])}"),
        ("Consumo comun", money(c["sharedCost"])),
        ("Exceso repartido", money(c["excess"])),
        ("Multas sin usar", money(c["surplus"])),
        ("Extras individuales", money(c["extrasTotal"])),
        ("TOTAL a recaudar", money(c["grand"])),
    ]
    for k, v in pairs:
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(70, 7, latin(k))
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, latin(v), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, latin("Cada quien paga"), new_x="LMARGIN", new_y="NEXT")
    with pdf.table(text_align=("LEFT", "CENTER", "RIGHT", "RIGHT", "RIGHT", "RIGHT"),
                   col_widths=(34, 16, 20, 18, 18, 22)) as table:
        table.row(["Persona", "Pold.", "Multas", "Exceso", "Extras", "Total"])
        for r in c["rows"]:
            table.row([
                latin(r["name"]), str(r["poldinas"]),
                money(r["fine"]),
                money(r["excessShare"]) if r["excessShare"] else "-",
                money(r["extras"]) if r["extras"] else "-",
                money(r["total"]),
            ])
        table.row(["TOTAL", "",
                   money(sum(r["fine"] for r in c["rows"])),
                   money(c["excess"]), money(c["extrasTotal"]), money(c["grand"])])
    pdf.ln(3)

    items = state.get("items", []) or []
    if items:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, latin("Factura"), new_x="LMARGIN", new_y="NEXT")
        with pdf.table(text_align=("LEFT", "CENTER", "RIGHT", "RIGHT", "CENTER", "LEFT")) as table:
            table.row(["Item", "Cant.", "Precio", "Subtotal", "Tipo", "Persona"])
            for i in items:
                kind = "Individual" if i.get("kind") == "individual" else "Comun"
                person = name_of.get(i.get("personId"), "-") if i.get("kind") == "individual" else "-"
                table.row([
                    latin(i.get("desc")), str(inum(i.get("qty"))),
                    money(num(i.get("unitPrice"))),
                    money(num(i.get("qty")) * num(i.get("unitPrice"))),
                    kind, latin(person),
                ])
        pdf.ln(3)

    fines = state.get("fines", []) or []
    if fines:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, latin("Poldinas registradas"), new_x="LMARGIN", new_y="NEXT")
        with pdf.table(text_align=("LEFT", "LEFT", "LEFT", "LEFT", "LEFT")) as table:
            table.row(["Persona", "Fecha", "Hora", "Motivo", "Descripcion"])
            for f in sorted(fines, key=lambda x: (x.get("date", ""), x.get("time", ""))):
                table.row([
                    latin(name_of.get(f.get("personId"), "-")), latin(f.get("date", "")),
                    latin(f.get("time", "")), latin(f.get("reason", "")), latin(f.get("note", "")),
                ])

    return bytes(pdf.output())


# =============================================================
#  Excel
# =============================================================
def build_xlsx(state):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    c = compute(state)
    event = state.get("event", {}) or {}
    name_of = {p.get("id"): (p.get("name") or "(sin nombre)") for p in state.get("people", [])}
    MONEY = "#,##0"
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="17352C")
    bold = Font(bold=True)

    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    rows = [
        ("Evento", event.get("name", "")),
        ("Lugar", event.get("place", "")),
        ("Fecha", event.get("date", "")),
        ("Valor poldina", c["pv"]),
        ("", ""),
        ("Total poldinas", c["totalPoldinas"]),
        ("Bolsa de multas", c["pool"]),
        ("Consumo comun", c["sharedCost"]),
        ("Exceso repartido", c["excess"]),
        ("Multas sin usar", c["surplus"]),
        ("Extras individuales", c["extrasTotal"]),
        ("TOTAL a recaudar", c["grand"]),
    ]
    for k, v in rows:
        ws.append([k, v])
    for cell in ws["B"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY
    ws["A12"].font = bold
    ws["B12"].font = bold
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22

    ws2 = wb.create_sheet("Cada quien paga")
    ws2.append(["Persona", "Poldinas", "Multas", "Exceso", "Extras", "Total"])
    for cell in ws2[1]:
        cell.font = head_font
        cell.fill = head_fill
    for r in c["rows"]:
        ws2.append([r["name"], r["poldinas"], r["fine"], r["excessShare"], r["extras"], r["total"]])
    ws2.append(["TOTAL", "", sum(r["fine"] for r in c["rows"]), c["excess"], c["extrasTotal"], c["grand"]])
    for col in ("C", "D", "E", "F"):
        for cell in ws2[col]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    for cell in ws2[ws2.max_row]:
        cell.font = bold
    ws2.column_dimensions["A"].width = 22
    for col in ("B", "C", "D", "E", "F"):
        ws2.column_dimensions[col].width = 13

    ws3 = wb.create_sheet("Factura")
    ws3.append(["Item", "Cantidad", "Precio unit", "Subtotal", "Tipo", "Persona"])
    for cell in ws3[1]:
        cell.font = head_font
        cell.fill = head_fill
    for i in (state.get("items", []) or []):
        q, p = num(i.get("qty")), num(i.get("unitPrice"))
        kind = "Individual" if i.get("kind") == "individual" else "Comun"
        person = name_of.get(i.get("personId"), "") if i.get("kind") == "individual" else ""
        ws3.append([i.get("desc", ""), inum(i.get("qty")), p, q * p, kind, person])
    for col in ("C", "D"):
        for cell in ws3[col]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    ws3.column_dimensions["A"].width = 26
    for col in ("B", "C", "D", "E", "F"):
        ws3.column_dimensions[col].width = 13

    ws4 = wb.create_sheet("Poldinas")
    ws4.append(["Persona", "Fecha", "Hora", "Motivo", "Descripcion"])
    for cell in ws4[1]:
        cell.font = head_font
        cell.fill = head_fill
    for f in (state.get("fines", []) or []):
        ws4.append([name_of.get(f.get("personId"), ""), f.get("date", ""), f.get("time", ""),
                    f.get("reason", ""), f.get("note", "")])
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 26
    ws4.column_dimensions["E"].width = 34

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def filename(state, ext):
    base = (state.get("event", {}) or {}).get("name") or "poldinas"
    safe = "".join(ch if (ch.isalnum() or ch in "-_") else "_" for ch in base).strip("_").lower() or "poldinas"
    return f"{safe}.{ext}"


# =============================================================
#  Rutas
# =============================================================
@app.route("/")
def home():
    return Response(PAGE, mimetype="text/html")


@app.route("/api/calc", methods=["POST"])
def api_calc():
    state = request.get_json(force=True, silent=True) or {}
    return jsonify(compute(state))


@app.route("/api/parse-invoice-text", methods=["POST"])
def api_parse_text():
    data = request.get_json(force=True, silent=True) or {}
    text = data.get("text", "")
    items = parse_invoice_text(text)
    meta = detect_totals_and_meta(text)
    return jsonify({"items": items, "meta": meta})


@app.route("/api/extract-invoice", methods=["POST"])
def api_extract_invoice():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no_file", "message": "No se recibió ningún archivo."}), 400

    mime = guess_mime(f)
    if mime not in ("image/jpeg", "image/png", "image/webp", "image/gif", "application/pdf"):
        return jsonify({"error": "bad_type",
                         "message": "Formato no soportado. Sube una imagen (jpg, png, webp) o un PDF."}), 200

    raw = f.read()
    if not raw:
        return jsonify({"error": "empty", "message": "El archivo llegó vacío."}), 200

    # PDF: primero intenta leer el texto real del archivo (gratis, sin OCR ni IA).
    # Solo funciona si el PDF trae texto de verdad, no si es una foto/escaneo.
    if mime == "application/pdf":
        try:
            pdf_text = extract_pdf_text(raw)
        except Exception:
            pdf_text = ""
        if pdf_text and len(pdf_text) >= 20:
            items = parse_invoice_text(pdf_text)
            if items:
                meta = detect_totals_and_meta(pdf_text)
                return jsonify({"items": items, "meta": meta})
        # si no se encontró texto util, sigue abajo e intenta con OCR local

    # Foto (o PDF escaneado sin texto real): intenta leer con OCR local, gratis y sin IA.
    ocr_text = ""
    ocr_failed = False
    try:
        if mime == "application/pdf":
            ocr_text = ocr_pdf_as_image(raw)
        else:
            ocr_text = ocr_image_text(raw)
    except OCRNotAvailable:
        ocr_failed = True
    except Exception:
        ocr_failed = True

    if ocr_text:
        items = parse_invoice_text(ocr_text)
        if items:
            meta = detect_totals_and_meta(ocr_text)
            return jsonify({"items": items, "meta": meta})

    try:
        parsed = extract_invoice_data(raw, mime)
    except MissingAPIKey:
        if ocr_failed:
            msg = ("El lector automático (OCR) no está disponible en este servidor todavía. "
                   "Usa 'Pegar texto' o agrega los items a mano mientras se ajusta.")
        else:
            msg = ("No se pudo reconocer texto útil en la imagen (puede estar muy inclinada, borrosa o "
                   "con poca luz). Prueba con una foto más nítida y derecha, usa 'Pegar texto' o agrega "
                   "los items a mano.")
        return jsonify({"error": "no_api_key", "message": msg}), 200
    except Exception as e:
        return jsonify({"error": "extract_failed", "message": f"No se pudo leer la factura: {e}"}), 200

    items = parsed.get("items", []) if isinstance(parsed, dict) else []
    norm = []
    for it in items:
        desc = str(it.get("desc", "")).strip() or "Item"
        kind_raw = str(it.get("kind", "")).lower()
        kind = "individual" if kind_raw.startswith("ind") else ("shared" if kind_raw else guess_kind(desc))
        norm.append({
            "desc": desc,
            "qty": max(1, inum(it.get("qty"), 1)),
            "unitPrice": round(num(it.get("unitPrice"), 0)),
            "kind": kind,
        })
    meta = {
        "place": parsed.get("place") if isinstance(parsed, dict) else None,
        "date": parsed.get("date") if isinstance(parsed, dict) else None,
        "subtotal": parsed.get("subtotal") if isinstance(parsed, dict) else None,
        "total": parsed.get("total") if isinstance(parsed, dict) else None,
    }
    return jsonify({"items": norm, "meta": meta})


@app.route("/api/export/pdf", methods=["POST"])
def api_pdf():
    state = request.get_json(force=True, silent=True) or {}
    try:
        data = build_pdf(state)
    except Exception as e:
        return jsonify({"error": f"No se pudo generar el PDF: {e}"}), 500
    return send_file(BytesIO(data), mimetype="application/pdf",
                      as_attachment=True, download_name=filename(state, "pdf"))


@app.route("/api/export/xlsx", methods=["POST"])
def api_xlsx():
    state = request.get_json(force=True, silent=True) or {}
    try:
        data = build_xlsx(state)
    except Exception as e:
        return jsonify({"error": f"No se pudo generar el Excel: {e}"}), 500
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename(state, "xlsx"))


# =============================================================
#  Interfaz (HTML + CSS + JS)
# =============================================================
PAGE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Poldinas - repartidor de la cuenta</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500;600;700&display=swap" rel="stylesheet">
<style>
  :root{
    --paper:#E7ECE2; --card:#FBFBF5; --card-2:#F2F4EC;
    --ink:#17352C; --ink-soft:#54655C; --line:#D4DACC;
    --brass:#A87A1E; --brass-deep:#815d13; --brass-soft:#F0E6CE;
    --brick:#C24E33; --brick-soft:#F4E1D9;
    --pine:#2A6A54; --pine-soft:#DBEBE1;
    --radius:14px; --shadow:0 1px 2px rgba(23,53,44,.06), 0 8px 24px rgba(23,53,44,.06);
  }
  *{box-sizing:border-box}
  html,body{margin:0}
  body{background:var(--paper); color:var(--ink); font-family:"Inter",system-ui,sans-serif; font-size:15px; line-height:1.5; -webkit-font-smoothing:antialiased;}
  .wrap{max-width:1120px; margin:0 auto; padding:26px 20px 80px;}
  header.top{display:flex; align-items:flex-end; gap:16px; flex-wrap:wrap; margin-bottom:22px;}
  .brand{display:flex; align-items:center; gap:14px;}
  .coin{width:52px; height:52px; border-radius:50%; flex:none;
    background:radial-gradient(circle at 34% 30%, #E7C878, var(--brass) 62%, var(--brass-deep));
    box-shadow:inset 0 2px 4px rgba(255,255,255,.5), inset 0 -3px 6px rgba(0,0,0,.25), var(--shadow);
    display:grid; place-items:center; color:#4a3608; font-family:"Fraunces",serif; font-weight:700; font-size:24px; border:2px solid #caa14e;}
  h1{font-family:"Fraunces",serif; font-weight:600; font-size:30px; line-height:1; margin:0; letter-spacing:-.01em;}
  .tag{color:var(--ink-soft); font-size:13px; margin-top:5px;}
  .top-actions{margin-left:auto; display:flex; gap:8px; flex-wrap:wrap;}
  .grid{display:grid; grid-template-columns:1fr 400px; gap:20px; align-items:start;}
  @media(max-width:900px){ .grid{grid-template-columns:1fr;} .side{position:static !important;} }
  .side{position:sticky; top:20px;}
  .card{background:var(--card); border:1px solid var(--line); border-radius:var(--radius); box-shadow:var(--shadow); margin-bottom:18px; overflow:hidden;}
  .card-h{display:flex; align-items:center; gap:10px; padding:15px 18px 13px; border-bottom:1px solid var(--line);}
  .card-h .n{font-family:"JetBrains Mono",monospace; font-size:12px; color:var(--brass-deep); background:var(--brass-soft); border-radius:6px; padding:2px 7px; font-weight:700;}
  .card-h h2{font-family:"Fraunces",serif; font-weight:600; font-size:18px; margin:0; flex:1;}
  .card-h .sub{color:var(--ink-soft); font-size:12.5px;}
  .card-b{padding:16px 18px;}
  label{display:block; font-size:12px; font-weight:600; color:var(--ink-soft); margin:0 0 5px;}
  input,select,textarea{width:100%; font-family:inherit; font-size:14px; color:var(--ink); background:#fff; border:1px solid var(--line); border-radius:9px; padding:9px 11px; outline:none;}
  textarea{resize:vertical; min-height:60px; font-family:inherit;}
  input:focus,select:focus,textarea:focus{border-color:var(--brass); box-shadow:0 0 0 3px var(--brass-soft);}
  input[type=number]{font-family:"JetBrains Mono",monospace;}
  .field-row{display:grid; gap:10px;}
  .cols-3{grid-template-columns:1.6fr 1fr auto;}
  @media(max-width:560px){ .cols-3{grid-template-columns:1fr;} }
  button{font-family:inherit; cursor:pointer; border:none; border-radius:9px; font-weight:600; font-size:13.5px; padding:9px 14px; transition:filter .12s, background .12s;}
  button:hover{filter:brightness(.97)}
  .btn-primary{background:var(--ink); color:#fff;}
  .btn-brass{background:var(--brass); color:#fff;}
  .btn-ghost{background:transparent; color:var(--ink); border:1px solid var(--line);}
  .btn-ghost:hover{background:var(--card-2)}
  .btn-sm{padding:7px 11px; font-size:12.5px;}
  .btn-x{background:transparent; color:var(--ink-soft); border:1px solid var(--line); padding:8px 10px; border-radius:8px; line-height:1;}
  .btn-x:hover{background:var(--brick-soft); color:var(--brick); border-color:var(--brick-soft);}
  .row{display:grid; gap:9px; align-items:center; padding:10px 0; border-top:1px dashed var(--line);}
  .row:first-child{border-top:none;}
  .row-people2{grid-template-columns:1fr 110px auto;}
  .row-item{grid-template-columns:1.3fr 60px 90px 100px 130px 90px auto;}
  @media(max-width:760px){ .row-item{grid-template-columns:1fr 1fr;} .row-people2{grid-template-columns:1fr auto auto;} }
  .row .lineamt{font-family:"JetBrains Mono",monospace; font-size:13px; text-align:right; color:var(--ink-soft); align-self:center; white-space:nowrap;}
  .empty{color:var(--ink-soft); font-size:13.5px; text-align:center; padding:16px 8px; font-style:italic;}
  .meter-wrap{margin:6px 0 14px;}
  .meter-title{font-size:11px; text-transform:uppercase; letter-spacing:.08em; color:var(--ink-soft); margin-bottom:7px; font-weight:600;}
  .meter{height:26px; border-radius:8px; background:var(--card-2); border:1px solid var(--line); display:flex; overflow:hidden;}
  .seg{height:100%; transition:width .3s ease;}
  .seg.covered{background:repeating-linear-gradient(45deg,var(--pine),var(--pine) 10px,#2f7a61 10px,#2f7a61 20px);}
  .seg.excess{background:repeating-linear-gradient(45deg,var(--brick),var(--brick) 8px,#d15b40 8px,#d15b40 16px);}
  .seg.surplus{background:repeating-linear-gradient(45deg,var(--brass-soft),var(--brass-soft) 8px,#e6d6a8 8px,#e6d6a8 16px);}
  .legend{display:flex; flex-wrap:wrap; gap:12px; margin-top:9px; font-size:11.5px; color:var(--ink-soft);}
  .legend span{display:inline-flex; align-items:center; gap:6px;}
  .dot{width:11px;height:11px;border-radius:3px; flex:none;}
  .dot.c{background:var(--pine)} .dot.e{background:var(--brick)} .dot.s{background:#e6d6a8}
  .stat-line{display:flex; justify-content:space-between; align-items:baseline; padding:7px 0; border-top:1px solid var(--line); font-size:13.5px;}
  .stat-line:first-of-type{border-top:none;}
  .stat-line .k{color:var(--ink-soft);}
  .stat-line .v{font-family:"JetBrains Mono",monospace; font-weight:600;}
  .stat-line.big{padding-top:12px;}
  .stat-line.big .k{font-family:"Fraunces",serif; font-size:16px; color:var(--ink);}
  .stat-line.big .v{font-size:19px; color:var(--brass-deep);}
  .note{font-size:12.5px; border-radius:10px; padding:10px 12px; margin-top:12px; line-height:1.45;}
  .note.warn{background:var(--brick-soft); color:#8a3620;}
  .note.info{background:var(--pine-soft); color:#1f4d3c;}
  table.tbl{width:100%; border-collapse:collapse; font-size:13px;}
  table.tbl th{text-align:right; font-size:11px; text-transform:uppercase; letter-spacing:.05em; color:var(--ink-soft); font-weight:600; padding:8px 8px; border-bottom:1px solid var(--line);}
  table.tbl th:first-child{text-align:left;}
  table.tbl td{padding:9px 8px; border-bottom:1px solid var(--line); text-align:right; font-family:"JetBrains Mono",monospace;}
  table.tbl td:first-child{text-align:left; font-family:"Inter",sans-serif; font-weight:500;}
  table.tbl tr.total-row td{border-top:2px solid var(--ink); border-bottom:none; font-weight:700; padding-top:11px;}
  table.tbl tr.total-row td:first-child{font-family:"Fraunces",serif;}
  .pill{display:inline-block; font-size:10.5px; padding:1px 7px; border-radius:20px; font-weight:600; margin-left:6px; vertical-align:middle; white-space:nowrap;}
  .pill.fined{background:var(--brass-soft); color:var(--brass-deep);}
  .pill.free{background:var(--pine-soft); color:var(--pine);}
  .grand{font-family:"Fraunces",serif; font-weight:700; color:var(--brass-deep);}
  .hist-item{display:flex; align-items:center; gap:10px; padding:9px 0; border-top:1px dashed var(--line); font-size:13px;}
  .hist-item:first-child{border-top:none;}
  .hist-item .hname{flex:1; font-weight:500;}
  .hist-item .hdate{color:var(--ink-soft); font-size:11.5px; font-family:"JetBrains Mono",monospace;}
  .toast{position:fixed; left:50%; bottom:26px; transform:translateX(-50%) translateY(20px); background:var(--ink); color:#fff; padding:11px 18px; border-radius:10px; font-size:13.5px; box-shadow:var(--shadow); opacity:0; transition:.25s; pointer-events:none; z-index:50; max-width:90vw; text-align:center;}
  .toast.show{opacity:1; transform:translateX(-50%) translateY(0);}
  .save-flag{font-size:11.5px; color:var(--ink-soft); margin-left:auto; display:flex; align-items:center; gap:6px;}
  .save-flag .d{width:7px;height:7px;border-radius:50%;background:var(--pine);}
  .dropzone{display:block; border:1.5px dashed var(--line); border-radius:10px; padding:16px; text-align:center; color:var(--ink-soft); font-size:13px; cursor:pointer; background:var(--card-2);}
  .dropzone:hover{border-color:var(--brass); color:var(--ink);}
  .dropzone input{display:none;}
  .divider{margin:12px 0; text-align:center; color:var(--ink-soft); font-size:11.5px; text-transform:uppercase; letter-spacing:.06em;}
  .fine-row{padding:10px 0; border-top:1px dashed var(--line);}
  .fine-row:first-child{border-top:none;}
  .fine-top{display:flex; align-items:center; gap:8px; flex-wrap:wrap;}
  .fine-name{font-weight:600;}
  .fine-reason{background:var(--brass-soft); color:var(--brass-deep); font-size:11.5px; padding:2px 8px; border-radius:20px;}
  .fine-date{font-family:"JetBrains Mono",monospace; font-size:11.5px; color:var(--ink-soft);}
  .fine-row .btn-x{margin-left:auto;}
  .fine-note{margin-top:4px; font-size:12.5px; color:var(--ink-soft); font-style:italic;}
  .staging-note{margin-bottom:10px;}
</style>
</head>
<body>
<div class="wrap">
  <header class="top">
    <div class="brand">
      <div class="coin">P</div>
      <div><h1>Poldinas</h1><div class="tag">Reparte la cuenta: multas, exceso y extras - cada quien lo suyo.</div></div>
    </div>
    <div class="top-actions">
      <button class="btn-ghost btn-sm" id="btnNew">Nuevo evento</button>
      <button class="btn-primary btn-sm" id="btnSave">Guardar evento</button>
    </div>
  </header>

  <div class="grid">
    <div class="main">
      <section class="card">
        <div class="card-h"><span class="n">01</span><h2>Evento</h2>
          <div class="save-flag"><span class="d"></span> <span id="saveState">Guardado</span></div></div>
        <div class="card-b">
          <div class="field-row cols-3">
            <div><label>Nombre / motivo</label><input id="evName" placeholder="Ej: Poldinas de julio"></div>
            <div><label>Lugar</label><input id="evPlace" placeholder="Ej: Pizzeria del centro"></div>
            <div style="width:130px"><label>Valor poldina</label><input id="evPV" type="number" min="0" step="100" value="1000"></div>
          </div>
          <div class="field-row" style="margin-top:10px; grid-template-columns:1fr; max-width:220px">
            <div><label>Fecha del consumo</label><input id="evDate" type="date"></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">02</span><h2>Personas</h2><span class="sub">La cantidad de poldinas se calcula sola</span></div>
        <div class="card-b">
          <div id="peopleList"></div>
          <div class="field-row" style="margin-top:14px; grid-template-columns:1fr auto; gap:10px">
            <div><label>Nombre</label><input id="pName" placeholder="Nombre de la persona"></div>
            <div style="display:flex; align-items:flex-end"><button class="btn-brass" id="btnAddPerson">Agregar</button></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">03</span><h2>Registrar poldina</h2><span class="sub">Fecha, hora y motivo de la multa</span></div>
        <div class="card-b">
          <div id="finesList"></div>
          <div class="field-row" style="margin-top:16px; grid-template-columns:1fr .8fr .7fr; gap:10px">
            <div><label>Persona</label><select id="fPerson"></select></div>
            <div><label>Fecha</label><input id="fDate" type="date"></div>
            <div><label>Hora</label><input id="fTime" type="time"></div>
          </div>
          <div class="field-row" style="margin-top:10px; grid-template-columns:1fr">
            <div>
              <label>Motivo</label>
              <select id="fReason">
                <option>Taza sucia o fuera de lugar</option>
                <option>Luz encendida</option>
                <option>Aire encendido</option>
                <option>Dejó el carnet</option>
                <option value="Otro">Otro</option>
              </select>
            </div>
          </div>
          <div class="field-row" id="fOtherWrap" style="display:none; margin-top:10px; grid-template-columns:1fr">
            <div><label>Especifica el motivo</label><input id="fReasonOther" placeholder="Escribe el motivo"></div>
          </div>
          <div class="field-row" style="margin-top:10px; grid-template-columns:1fr">
            <div><label>Descripción (opcional)</label><textarea id="fNote" placeholder="Detalles adicionales..."></textarea></div>
          </div>
          <button class="btn-brass" id="btnAddFine" style="margin-top:12px; width:100%">Registrar poldina</button>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">04</span><h2>Factura</h2><span class="sub">Lo común se paga con multas, lo individual aparte</span></div>
        <div class="card-b">
          <label class="dropzone" for="invoiceFile">
            Sube una foto o PDF de la factura para leerla automáticamente
            <input type="file" id="invoiceFile" accept="image/*,application/pdf">
          </label>
          <div style="font-size:11.5px; color:var(--ink-soft); margin-top:6px;">
            Los PDF con texto real se leen sin costo. Las fotos se intentan leer con OCR local (gratis, sin IA); si la foto está muy inclinada o borrosa puede no reconocer nada.
          </div>
          <div class="divider">— o —</div>
          <label>Pega el texto de la factura</label>
          <textarea id="invoiceText" placeholder="Ej:&#10;2 Pizza mediana 24000&#10;Gaseosa 3000&#10;Cerveza 5000"></textarea>
          <button class="btn-ghost btn-sm" id="btnParseText" style="margin-top:8px">Leer texto</button>

          <div id="stagingBox" style="display:none; margin-top:16px;"></div>

          <div style="margin:20px 0 8px; font-weight:600; font-size:13px; color:var(--ink-soft);">Items confirmados</div>
          <div id="itemsList"></div>
          <div class="field-row" style="margin-top:14px; grid-template-columns:1.3fr .6fr .9fr .9fr 1fr auto; gap:10px" id="addItemRow">
            <div><label>Descripción</label><input id="iDesc" placeholder="Ej: Pizza mediana"></div>
            <div><label>Cant.</label><input id="iQty" type="number" min="0" step="1" value="1"></div>
            <div><label>Precio unit.</label><input id="iPrice" type="number" min="0" step="100" placeholder="1000"></div>
            <div><label>Tipo</label>
              <select id="iKind"><option value="shared">Común</option><option value="individual">Individual</option></select>
            </div>
            <div id="iPersonWrap" style="display:none"><label>Persona</label><select id="iPerson"></select></div>
            <div style="display:flex; align-items:flex-end"><button class="btn-brass" id="btnAddItem" style="width:100%">Agregar</button></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">05</span><h2>Eventos guardados</h2></div>
        <div class="card-b"><div id="histList"></div></div>
      </section>
    </div>

    <div class="side">
      <section class="card">
        <div class="card-h"><h2 style="font-size:20px">Resumen</h2></div>
        <div class="card-b" style="padding:18px">
          <div class="meter-wrap">
            <div class="meter-title">Bolsa de multas vs. consumo comun</div>
            <div class="meter" id="meter"></div>
            <div class="legend">
              <span><i class="dot c"></i>Cubierto por multas</span>
              <span><i class="dot e"></i>Exceso (sin multa)</span>
              <span><i class="dot s"></i>Multas sin usar</span>
            </div>
          </div>
          <div class="stat-line"><span class="k">Bolsa de multas <span id="poldCount" class="mono"></span></span><span class="v" id="vPool">$0</span></div>
          <div class="stat-line"><span class="k">Consumo comun</span><span class="v" id="vShared">$0</span></div>
          <div class="stat-line"><span class="k">Exceso a repartir</span><span class="v" id="vExcess" style="color:var(--brick)">$0</span></div>
          <div class="stat-line"><span class="k">Extras individuales</span><span class="v" id="vExtras">$0</span></div>
          <div class="stat-line big"><span class="k">Total a recaudar</span><span class="v" id="vGrand">$0</span></div>
          <div id="noteBox"></div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><h2 style="font-size:18px">Cada quien paga</h2></div>
        <div class="card-b" style="padding:10px 12px 14px">
          <div id="breakdown"></div>
          <div style="display:flex; gap:8px; margin-top:16px;">
            <button class="btn-primary" id="btnPDF" style="flex:1">Exportar PDF</button>
            <button class="btn-brass" id="btnXLS" style="flex:1">Exportar Excel</button>
          </div>
        </div>
      </section>
    </div>
  </div>
</div>
<div class="toast" id="toast"></div>

<script>
const uid = () => Date.now().toString(36) + Math.random().toString(36).slice(2,7);
const CUR = new Intl.NumberFormat('es-DO',{style:'currency',currency:'DOP',maximumFractionDigits:0});
const money = n => CUR.format(Math.round(n||0));

const LS_CUR='poldinas:current', LS_EVENTS='poldinas:events';
let state = loadCurrent() || blank();
let staging = [];

function blank(){ return {id:uid(), event:{name:'',place:'',date:'',poldinaValue:1000}, people:[], fines:[], items:[]}; }
function loadCurrent(){ try{ const r=localStorage.getItem(LS_CUR); const s=r?JSON.parse(r):null;
  if(s && !s.fines) s.fines=[]; if(s && !s.items) s.items=[]; return s; }catch(e){ return null; } }
function loadEvents(){ try{ const r=localStorage.getItem(LS_EVENTS); return r?JSON.parse(r):[]; }catch(e){ return []; } }

let saveT=null;
function save(){ const s=document.getElementById('saveState'); if(s) s.textContent='Guardando...';
  clearTimeout(saveT); saveT=setTimeout(()=>{ try{ localStorage.setItem(LS_CUR, JSON.stringify(state)); if(s) s.textContent='Guardado'; }catch(e){ if(s) s.textContent='No se pudo guardar'; } }, 300); }

/* ---------- calculo: pide a Python, con respaldo local ---------- */
let calcT=null;
function scheduleCalc(){ clearTimeout(calcT); calcT=setTimeout(doCalc, 250); }
async function doCalc(){
  try{
    const res = await fetch('/api/calc',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(state)});
    if(!res.ok) throw 0;
    renderSummary(await res.json());
  }catch(e){ renderSummary(jsCalc()); }
}
function jsCalc(){
  const pv=+state.event.poldinaValue||1000;
  const finesByPerson={};
  state.fines.forEach(f=>{ finesByPerson[f.personId]=(finesByPerson[f.personId]||0)+1; });
  const totalPold=Object.values(finesByPerson).reduce((a,b)=>a+b,0);
  const pool=totalPold*pv;
  const sharedItems=state.items.filter(i=>i.kind!=='individual');
  const indivItems=state.items.filter(i=>i.kind==='individual');
  const indivAssigned=indivItems.filter(i=>i.personId);
  const indivUnassigned=indivItems.filter(i=>!i.personId);
  const sharedCost=sharedItems.reduce((s,i)=>s+(+i.qty||0)*(+i.unitPrice||0),0);
  const excess=Math.max(0,sharedCost-pool), surplus=Math.max(0,pool-sharedCost), covered=Math.min(sharedCost,pool);
  const per = state.people.length ? excess/state.people.length : 0;
  const rows=state.people.map(p=>{ const pold=finesByPerson[p.id]||0; const fine=pold*pv;
    const ex=indivAssigned.filter(i=>i.personId===p.id).reduce((s,i)=>s+(+i.qty||0)*(+i.unitPrice||0),0);
    const sh=excess>0 ? per : 0;
    return {id:p.id,name:p.name||'(sin nombre)',poldinas:pold,fine,excessShare:sh,extras:ex,total:fine+ex+sh,fined:pold>0}; });
  const extrasTotal=indivAssigned.reduce((s,i)=>s+(+i.qty||0)*(+i.unitPrice||0),0);
  const unassignedTotal=indivUnassigned.reduce((s,i)=>s+(+i.qty||0)*(+i.unitPrice||0),0);
  return {totalPoldinas:totalPold,pool,sharedCost,covered,excess,surplus,extrasTotal,unassignedTotal,unassignedCount:indivUnassigned.length,grand:rows.reduce((s,r)=>s+r.total,0),excessPer:per,rows};
}

/* ---------- helpers DOM ---------- */
function el(tag, attrs={}, kids=[]){ const n=document.createElement(tag);
  for(const k in attrs){ if(k==='class')n.className=attrs[k]; else if(k==='html')n.innerHTML=attrs[k];
    else if(k.startsWith('on'))n.addEventListener(k.slice(2),attrs[k]); else n.setAttribute(k,attrs[k]); }
  (Array.isArray(kids)?kids:[kids]).forEach(c=>{ if(c!=null)n.append(c.nodeType?c:document.createTextNode(c)); });
  return n; }
function fillOptions(sel, selected){ sel.innerHTML='';
  if(!state.people.length){ sel.append(el('option',{value:''},'- agrega personas -')); return; }
  state.people.forEach(p=>{ const o=el('option',{value:p.id}, p.name||'(sin nombre)'); if(p.id===selected)o.selected=true; sel.append(o); }); }
function kindSelect(value, onChange){
  const sel=document.createElement('select');
  [['shared','Común'],['individual','Individual']].forEach(([v,label])=>{
    const o=document.createElement('option'); o.value=v; o.textContent=label; if(v===value)o.selected=true; sel.appendChild(o);
  });
  sel.addEventListener('change', onChange);
  return sel;
}
function nameOf(id){ return (state.people.find(p=>p.id===id)||{}).name || '(sin nombre)'; }
function refreshPersonSelects(){
  fillOptions(document.getElementById('fPerson'));
  fillOptions(document.getElementById('iPerson'));
  renderFines(); renderItems(); renderStaging();
}

/* ---------- personas ---------- */
function renderPeople(){ const box=document.getElementById('peopleList'); box.innerHTML='';
  if(!state.people.length){ box.append(el('div',{class:'empty'},'Aun no hay personas.')); return; }
  state.people.forEach(p=>{
    const count=state.fines.filter(f=>f.personId===p.id).length;
    const name=el('input',{value:p.name||'',placeholder:'Nombre',oninput:e=>{p.name=e.target.value; refreshPersonSelects(); renderPeople.keepFocus=true; scheduleCalc(); save();}});
    const badge=el('span',{class: count?'pill fined':'pill free'}, count? (count+' poldina'+(count>1?'s':'')) : 'sin multa');
    const del=el('button',{class:'btn-x',title:'Quitar',onclick:()=>{
      state.people=state.people.filter(x=>x.id!==p.id);
      state.fines=state.fines.filter(x=>x.personId!==p.id);
      state.items=state.items.filter(x=>!(x.kind==='individual'&&x.personId===p.id));
      renderPeople(); refreshPersonSelects(); scheduleCalc(); save();
    }},'X');
    box.append(el('div',{class:'row row-people2'},[name,badge,del]));
  });
}

/* ---------- poldinas / multas ---------- */
function renderFines(){ const box=document.getElementById('finesList'); box.innerHTML='';
  if(!state.fines.length){ box.append(el('div',{class:'empty'},'Sin poldinas registradas.')); return; }
  const sorted=[...state.fines].sort((a,b)=> (b.date+' '+b.time).localeCompare(a.date+' '+a.time));
  sorted.forEach(f=>{
    const del=el('button',{class:'btn-x',title:'Quitar',onclick:()=>{
      state.fines=state.fines.filter(x=>x.id!==f.id); renderFines(); renderPeople(); scheduleCalc(); save();
    }},'X');
    const meta=(f.date||'')+' '+(f.time||'');
    const top=el('div',{class:'fine-top'},[
      el('span',{class:'fine-name'}, nameOf(f.personId)),
      el('span',{class:'fine-reason'}, f.reason||''),
      el('span',{class:'fine-date'}, meta.trim()),
      del,
    ]);
    const wrap=el('div',{class:'fine-row'},[top]);
    if(f.note) wrap.append(el('div',{class:'fine-note'}, f.note));
    box.append(wrap);
  });
}

/* ---------- factura: items confirmados ---------- */
function renderItems(){ const box=document.getElementById('itemsList'); box.innerHTML='';
  if(!state.items.length){ box.append(el('div',{class:'empty'},'Sin items en la factura todavia.')); return; }
  state.items.forEach(it=>{
    const desc=el('input',{value:it.desc||'',oninput:e=>{it.desc=e.target.value; save();}});
    const qty=el('input',{type:'number',min:'0',step:'1',value:it.qty??1,oninput:e=>{it.qty=+e.target.value||0; upd(); scheduleCalc(); save();}});
    const price=el('input',{type:'number',min:'0',step:'100',value:it.unitPrice??0,oninput:e=>{it.unitPrice=+e.target.value||0; upd(); scheduleCalc(); save();}});
    const kSel=kindSelect(it.kind, e=>{ it.kind=e.target.value; if(it.kind==='shared') it.personId=null; renderItems(); scheduleCalc(); save(); });
    const pSel=document.createElement('select'); fillOptions(pSel, it.personId);
    pSel.style.display = it.kind==='individual' ? '' : 'none';
    pSel.addEventListener('change', e=>{ it.personId=e.target.value; scheduleCalc(); save(); });
    const amt=el('div',{class:'lineamt'}, money((+it.qty||0)*(+it.unitPrice||0)));
    function upd(){ amt.textContent=money((+it.qty||0)*(+it.unitPrice||0)); }
    const del=el('button',{class:'btn-x',title:'Quitar',onclick:()=>{ state.items=state.items.filter(x=>x.id!==it.id); renderItems(); scheduleCalc(); save(); }},'X');
    box.append(el('div',{class:'row row-item'},[desc,qty,price,kSel,pSel,amt,del]));
  });
}

/* ---------- factura: staging (pendiente de confirmar) ---------- */
function renderStaging(){ const box=document.getElementById('stagingBox'); box.innerHTML='';
  if(!staging.length){ box.style.display='none'; return; }
  box.style.display='block';
  box.append(el('div',{class:'note info staging-note'}, staging.length+' item(s) leidos de la factura. Revisa cantidad, precio y tipo antes de confirmar.'));
  staging.forEach((it,idx)=>{
    const desc=el('input',{value:it.desc||'',oninput:e=>{it.desc=e.target.value;}});
    const qty=el('input',{type:'number',min:'0',step:'1',value:it.qty??1,oninput:e=>{it.qty=+e.target.value||0; upd();}});
    const price=el('input',{type:'number',min:'0',step:'100',value:it.unitPrice??0,oninput:e=>{it.unitPrice=+e.target.value||0; upd();}});
    const kSel=kindSelect(it.kind, e=>{ it.kind=e.target.value; renderStaging(); });
    const pSel=document.createElement('select'); fillOptions(pSel, it.personId);
    pSel.style.display = it.kind==='individual' ? '' : 'none';
    pSel.addEventListener('change', e=>{ it.personId=e.target.value; });
    const amt=el('div',{class:'lineamt'}, money((+it.qty||0)*(+it.unitPrice||0)));
    function upd(){ amt.textContent=money((+it.qty||0)*(+it.unitPrice||0)); }
    const del=el('button',{class:'btn-x',onclick:()=>{ staging.splice(idx,1); renderStaging(); }},'X');
    box.append(el('div',{class:'row row-item'},[desc,qty,price,kSel,pSel,amt,del]));
  });
  const confirmBtn=el('button',{class:'btn-primary',style:'margin-top:12px',onclick:confirmStaging},'Agregar todo a la factura');
  const discardBtn=el('button',{class:'btn-ghost',style:'margin-top:12px; margin-left:8px',onclick:()=>{ staging=[]; renderStaging(); }},'Descartar');
  box.append(confirmBtn, discardBtn);
}
function confirmStaging(){
  staging.forEach(it=>{
    state.items.push({id:uid(), desc:it.desc||'Item', qty:+it.qty||0, unitPrice:+it.unitPrice||0,
      kind: it.kind==='individual'?'individual':'shared',
      personId: it.kind==='individual' ? (it.personId||null) : null});
  });
  staging=[]; renderStaging(); renderItems(); scheduleCalc(); save();
  toast('Items agregados a la factura');
}

/* ---------- resumen ---------- */
function renderSummary(c){ renderMeter(c); renderStats(c); renderBreakdown(c); }
function renderMeter(c){ const m=document.getElementById('meter'); m.innerHTML='';
  const scale=Math.max(c.pool,c.sharedCost,1), w=v=>Math.max(0,(v/scale)*100);
  m.append(el('div',{class:'seg covered',style:`width:${w(c.covered)}%`,title:'Cubierto: '+money(c.covered)}));
  if(c.excess>0) m.append(el('div',{class:'seg excess',style:`width:${w(c.excess)}%`,title:'Exceso: '+money(c.excess)}));
  if(c.surplus>0) m.append(el('div',{class:'seg surplus',style:`width:${w(c.surplus)}%`,title:'Sin usar: '+money(c.surplus)}));
}
function renderStats(c){
  document.getElementById('poldCount').textContent='('+c.totalPoldinas+' pold.)';
  document.getElementById('vPool').textContent=money(c.pool);
  document.getElementById('vShared').textContent=money(c.sharedCost);
  document.getElementById('vExcess').textContent=money(c.excess);
  document.getElementById('vExtras').textContent=money(c.extrasTotal);
  document.getElementById('vGrand').textContent=money(c.grand);
  const nb=document.getElementById('noteBox'); nb.innerHTML='';
  if(c.excess>0) nb.append(el('div',{class:'note warn'},`El consumo se paso por ${money(c.excess)}. Se reparte entre las ${state.people.length} persona(s) -> ${money(c.excessPer)} c/u.`));
  if(c.surplus>0) nb.append(el('div',{class:'note info'},`Las multas superan el consumo comun por ${money(c.surplus)} (bolsa sin usar).`));
  if(c.unassignedCount>0) nb.append(el('div',{class:'note warn'},`Hay ${money(c.unassignedTotal)} en ${c.unassignedCount} item(s) individuales sin asignar a nadie. Ve a "Factura" y elige quien lo consumio, o no se cobrara.`));
}
function renderBreakdown(c){ const box=document.getElementById('breakdown'); box.innerHTML='';
  if(!c.rows.length){ box.append(el('div',{class:'empty'},'Agrega personas para ver el reparto.')); return; }
  const tbl=el('table',{class:'tbl'});
  tbl.append(el('thead',{},el('tr',{},[el('th',{},'Persona'),el('th',{},'Multas'),el('th',{},'Exceso'),el('th',{},'Extras'),el('th',{},'Total')])));
  const tb=el('tbody');
  c.rows.forEach(r=>{
    const pill = r.fined ? el('span',{class:'pill fined'}, r.poldinas+'p') : el('span',{class:'pill free'},'sin multa');
    tb.append(el('tr',{},[ el('td',{},[document.createTextNode(r.name),pill]),
      el('td',{},money(r.fine)), el('td',{}, r.excessShare?money(r.excessShare):'-'),
      el('td',{}, r.extras?money(r.extras):'-'), el('td',{class:'grand'},money(r.total)) ]));
  });
  tb.append(el('tr',{class:'total-row'},[ el('td',{},'Total'),
    el('td',{},money(c.rows.reduce((s,r)=>s+r.fine,0))), el('td',{},money(c.excess)),
    el('td',{},money(c.extrasTotal)), el('td',{class:'grand'},money(c.grand)) ]));
  tbl.append(tb); box.append(tbl);
}

/* ---------- historial ---------- */
function renderHistory(){ const box=document.getElementById('histList'); box.innerHTML='';
  const evs=loadEvents().sort((a,b)=>(b.savedAt||'').localeCompare(a.savedAt||''));
  if(!evs.length){ box.append(el('div',{class:'empty'},'Todavia no has guardado eventos.')); return; }
  evs.forEach(ev=>{
    const d=ev.savedAt?new Date(ev.savedAt).toLocaleDateString('es-DO',{day:'2-digit',month:'short',year:'2-digit'}):'';
    const load=el('button',{class:'btn-ghost btn-sm',onclick:()=>{ state=JSON.parse(JSON.stringify(ev.state)); if(!state.fines)state.fines=[]; if(!state.items)state.items=[]; syncInputs(); renderAll(); save(); toast('Evento cargado'); }},'Cargar');
    const del=el('button',{class:'btn-x',title:'Eliminar',onclick:()=>{ const list=loadEvents().filter(e=>e.id!==ev.id); localStorage.setItem(LS_EVENTS,JSON.stringify(list)); renderHistory(); toast('Evento eliminado'); }},'X');
    box.append(el('div',{class:'hist-item'},[ el('span',{class:'hname'}, ev.name||'Evento'), el('span',{class:'hdate'}, d), load, del ]));
  });
}

/* ---------- inputs evento ---------- */
function syncInputs(){
  document.getElementById('evName').value=state.event.name||'';
  document.getElementById('evPlace').value=state.event.place||'';
  document.getElementById('evDate').value=state.event.date||'';
  document.getElementById('evPV').value=state.event.poldinaValue??1000;
}
function renderAll(){ renderPeople(); renderFines(); renderItems(); refreshPersonSelects(); scheduleCalc(); }

function applyInvoiceMeta(meta){
  if(!meta) return;
  if(meta.place && !state.event.place){
    state.event.place = meta.place; document.getElementById('evPlace').value = meta.place;
  }
  if(meta.date && !state.event.date){
    state.event.date = meta.date; document.getElementById('evDate').value = meta.date;
  }
  if(!state.event.name && meta.place){
    state.event.name = 'Consumo en ' + meta.place; document.getElementById('evName').value = state.event.name;
  }
  save();
}
function addTaxDifferenceIfNeeded(meta){
  if(!meta || !meta.total) return;
  const sumStaging = staging.reduce((s,it)=> s+(+it.qty||0)*(+it.unitPrice||0), 0);
  const diff = Math.round((+meta.total||0) - sumStaging);
  if(diff > 1){
    staging.push({desc:'Cargos e impuestos (ITBIS/Ley/servicio)', qty:1, unitPrice:diff, kind:'shared', personId:null});
    toast('Se detecto un total de '+money(meta.total)+'. Se agrego la diferencia ('+money(diff)+') como cargo comun.');
  }
}

/* ---------- factura: subir archivo / pegar texto ---------- */
async function handleInvoiceFile(file){
  toast('Leyendo factura...');
  const fd=new FormData(); fd.append('file', file);
  try{
    const res=await fetch('/api/extract-invoice',{method:'POST', body:fd});
    const data=await res.json();
    if(data.error==='no_api_key'){ toast(data.message||'Extraccion automatica no configurada.'); return; }
    if(data.error){ toast(data.message||'No se pudo leer la factura.'); return; }
    if(!data.items || !data.items.length){ toast('No se encontraron items en la factura.'); return; }
    staging = data.items.map(it=>({...it, personId:null}));
    addTaxDifferenceIfNeeded(data.meta);
    applyInvoiceMeta(data.meta);
    renderStaging(); toast(data.items.length+' item(s) leidos, revisalos abajo');
  }catch(e){ toast('No se pudo leer la factura (revisa la conexion)'); }
}
async function handleInvoiceText(){
  const text=document.getElementById('invoiceText').value.trim();
  if(!text){ toast('Pega el texto de la factura primero'); return; }
  try{
    const res=await fetch('/api/parse-invoice-text',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({text})});
    const data=await res.json();
    if(!data.items || !data.items.length){ toast('No se reconocieron lineas. Revisa el formato o agrega a mano.'); return; }
    staging = data.items.map(it=>({...it, personId:null}));
    addTaxDifferenceIfNeeded(data.meta);
    applyInvoiceMeta(data.meta);
    renderStaging(); toast(data.items.length+' linea(s) reconocidas, revisalas abajo');
  }catch(e){ toast('No se pudo leer el texto'); }
}

/* ---------- exportar ---------- */
async function download(url){
  try{
    const res=await fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(state)});
    if(!res.ok) throw 0;
    const blob=await res.blob(); const dl=URL.createObjectURL(blob);
    const cd=res.headers.get('Content-Disposition')||''; const m=cd.match(/filename="?([^"]+)"?/);
    const a=document.createElement('a'); a.href=dl; a.download=m?m[1]:'poldinas'; document.body.append(a); a.click(); a.remove(); URL.revokeObjectURL(dl);
  }catch(e){ toast('No se pudo exportar (revisa la conexion)'); }
}

/* ---------- toast ---------- */
let tT=null; function toast(msg){ const t=document.getElementById('toast'); t.textContent=msg; t.classList.add('show'); clearTimeout(tT); tT=setTimeout(()=>t.classList.remove('show'),2600); }

/* ---------- bind ---------- */
function bind(){ const on=(id,ev,fn)=>document.getElementById(id).addEventListener(ev,fn);
  on('evName','input',e=>{state.event.name=e.target.value; save();});
  on('evPlace','input',e=>{state.event.place=e.target.value; save();});
  on('evDate','input',e=>{state.event.date=e.target.value; save();});
  on('evPV','input',e=>{state.event.poldinaValue=+e.target.value||0; scheduleCalc(); save();});

  on('btnAddPerson','click',()=>{ const n=document.getElementById('pName').value.trim(); if(!n){toast('Escribe un nombre');return;}
    state.people.push({id:uid(),name:n}); document.getElementById('pName').value='';
    renderPeople(); refreshPersonSelects(); scheduleCalc(); save(); document.getElementById('pName').focus(); });

  document.getElementById('fReason').addEventListener('change', e=>{
    document.getElementById('fOtherWrap').style.display = e.target.value==='Otro' ? '' : 'none'; });

  on('btnAddFine','click',()=>{
    if(!state.people.length){ toast('Primero agrega personas'); return; }
    const personId=document.getElementById('fPerson').value; if(!personId){ toast('Selecciona una persona'); return; }
    const reasonSel=document.getElementById('fReason').value;
    const reason = reasonSel==='Otro' ? (document.getElementById('fReasonOther').value.trim()||'Otro') : reasonSel;
    const note=document.getElementById('fNote').value.trim();
    state.fines.push({id:uid(), personId,
      date: document.getElementById('fDate').value || new Date().toISOString().slice(0,10),
      time: document.getElementById('fTime').value || '',
      reason, note});
    document.getElementById('fReasonOther').value=''; document.getElementById('fNote').value='';
    renderFines(); renderPeople(); scheduleCalc(); save(); toast('Poldina registrada');
  });

  document.getElementById('iKind').addEventListener('change', e=>{
    document.getElementById('iPersonWrap').style.display = e.target.value==='individual' ? '' : 'none'; });

  on('btnAddItem','click',()=>{
    const d=document.getElementById('iDesc').value.trim(); if(!d){toast('Escribe la descripcion');return;}
    const kind=document.getElementById('iKind').value;
    const personId = kind==='individual' ? document.getElementById('iPerson').value : null;
    if(kind==='individual' && !personId){ toast('Selecciona quien lo consumio'); return; }
    state.items.push({id:uid(), desc:d, qty:+document.getElementById('iQty').value||0,
      unitPrice:+document.getElementById('iPrice').value||0, kind, personId});
    document.getElementById('iDesc').value=''; document.getElementById('iQty').value=1; document.getElementById('iPrice').value='';
    renderItems(); scheduleCalc(); save(); document.getElementById('iDesc').focus();
  });

  document.getElementById('invoiceFile').addEventListener('change', e=>{
    const f=e.target.files[0]; if(f) handleInvoiceFile(f); e.target.value='';
  });
  on('btnParseText','click', handleInvoiceText);

  on('btnSave','click',()=>{ const list=loadEvents(); const snap={id:state.id,name:state.event.name||'Evento sin nombre',savedAt:new Date().toISOString(),state:JSON.parse(JSON.stringify(state))};
    const i=list.findIndex(e=>e.id===state.id); if(i>=0)list[i]=snap; else list.push(snap);
    localStorage.setItem(LS_EVENTS,JSON.stringify(list)); renderHistory(); toast('Evento guardado'); });
  on('btnNew','click',()=>{ state=blank(); syncInputs(); renderAll(); save(); toast('Evento nuevo'); });
  on('btnPDF','click',()=>download('/api/export/pdf'));
  on('btnXLS','click',()=>download('/api/export/xlsx'));
}

/* ---------- init ---------- */
bind(); syncInputs(); renderAll(); renderHistory();
(function setDefaultDateTime(){
  const d=new Date();
  document.getElementById('fDate').value = d.toISOString().slice(0,10);
  document.getElementById('fTime').value = d.toTimeString().slice(0,5);
})();
</script>
</body>
</html>"""

if __name__ == "__main__":
    app.run(debug=True, port=5000)
