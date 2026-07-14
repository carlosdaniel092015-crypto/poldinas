from io import BytesIO

from flask import Flask, request, jsonify, send_file, Response

app = Flask(__name__)


# =============================================================
#  Utilidades
# =============================================================
def num(v, d=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def inum(v, d=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return d


def money(n):
    """Formato pesos colombianos: $ 24.000"""
    s = f"{int(round(n or 0)):,}".replace(",", ".")
    return f"$ {s}"


def latin(s):
    """Deja el texto seguro para las fuentes base del PDF."""
    return str(s or "").encode("latin-1", "replace").decode("latin-1")


# =============================================================
#  Cálculo (fuente única de verdad)
# =============================================================
def compute(state):
    event = state.get("event", {}) or {}
    pv = num(event.get("poldinaValue"), 1000) or 1000
    people = state.get("people", []) or []
    shared = state.get("shared", []) or []
    extras = state.get("extras", []) or []

    total_pold = sum(inum(p.get("poldinas")) for p in people)
    pool = total_pold * pv
    shared_cost = sum(num(i.get("qty")) * num(i.get("unitPrice")) for i in shared)
    excess = max(0.0, shared_cost - pool)
    surplus = max(0.0, pool - shared_cost)
    covered = min(shared_cost, pool)

    non_fined = [p for p in people if inum(p.get("poldinas")) == 0]
    base = non_fined
    fallback = False
    if excess > 0 and len(non_fined) == 0:
        base = people
        fallback = True
    excess_per = (excess / len(base)) if base else 0.0
    base_ids = {p.get("id") for p in base}

    rows = []
    for p in people:
        pold = inum(p.get("poldinas"))
        fine = pold * pv
        ex = sum(num(e.get("amount")) for e in extras if e.get("personId") == p.get("id"))
        pays = excess > 0 and p.get("id") in base_ids
        eshare = excess_per if pays else 0.0
        rows.append({
            "id": p.get("id"),
            "name": p.get("name") or "(sin nombre)",
            "poldinas": pold,
            "fine": fine,
            "excessShare": eshare,
            "extras": ex,
            "total": fine + eshare + ex,
            "fined": pold > 0,
        })

    extras_total = sum(num(e.get("amount")) for e in extras)
    grand = sum(r["total"] for r in rows)

    return {
        "pv": pv,
        "totalPoldinas": total_pold,
        "pool": pool,
        "sharedCost": shared_cost,
        "covered": covered,
        "excess": excess,
        "surplus": surplus,
        "extrasTotal": extras_total,
        "grand": grand,
        "excessPer": excess_per,
        "nonFinedCount": len(non_fined),
        "fallback": fallback,
        "rows": rows,
    }


# =============================================================
#  PDF
# =============================================================
def build_pdf(state):
    from fpdf import FPDF

    c = compute(state)
    event = state.get("event", {}) or {}

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(23, 53, 44)
    pdf.cell(0, 10, latin("Poldinas - reparto de la cuenta"), new_x="LMARGIN", new_y="NEXT")

    meta = "   |   ".join([x for x in [event.get("name"), event.get("place"), event.get("date")] if x])
    if meta:
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(84, 101, 92)
        pdf.cell(0, 7, latin(meta), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Resumen
    pdf.set_text_color(23, 53, 44)
    pairs = [
        ("Bolsa de multas", f"{c['totalPoldinas']} poldinas = {money(c['pool'])}"),
        ("Consumo comun", money(c["sharedCost"])),
        ("Exceso repartido", money(c["excess"])),
        ("Multas sin usar", money(c["surplus"])),
        ("Extras individuales", money(c["extrasTotal"])),
        ("TOTAL a recaudar", money(c["grand"])),
    ]
    for k, v in pairs:
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(70, 7, latin(k))
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, latin(v), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Reparto por persona
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, latin("Cada quien paga"), new_x="LMARGIN", new_y="NEXT")
    with pdf.table(text_align=("LEFT", "CENTER", "RIGHT", "RIGHT", "RIGHT", "RIGHT"),
                   col_widths=(34, 16, 20, 18, 18, 22)) as table:
        table.row(["Persona", "Pold.", "Multas", "Exceso", "Extras", "Total"])
        for r in c["rows"]:
            table.row([
                latin(r["name"]), str(r["poldinas"]),
                money(r["fine"]),
                money(r["excessShare"]) if r["excessShare"] else "-",
                money(r["extras"]) if r["extras"] else "-",
                money(r["total"]),
            ])
        table.row(["TOTAL", "",
                   money(sum(r["fine"] for r in c["rows"])),
                   money(c["excess"]), money(c["extrasTotal"]), money(c["grand"])])
    pdf.ln(3)

    shared = state.get("shared", []) or []
    if shared:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, latin("Consumo comun"), new_x="LMARGIN", new_y="NEXT")
        with pdf.table(text_align=("LEFT", "CENTER", "RIGHT", "RIGHT")) as table:
            table.row(["Item", "Cant.", "Precio", "Subtotal"])
            for i in shared:
                table.row([latin(i.get("desc")), str(inum(i.get("qty"))),
                           money(num(i.get("unitPrice"))),
                           money(num(i.get("qty")) * num(i.get("unitPrice")))])
        pdf.ln(3)

    extras = state.get("extras", []) or []
    if extras:
        name_of = {p.get("id"): (p.get("name") or "(sin nombre)") for p in state.get("people", [])}
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, latin("Consumos individuales"), new_x="LMARGIN", new_y="NEXT")
        with pdf.table(text_align=("LEFT", "LEFT", "RIGHT")) as table:
            table.row(["Persona", "Concepto", "Monto"])
            for e in extras:
                table.row([latin(name_of.get(e.get("personId"), "(sin asignar)")),
                           latin(e.get("desc")), money(num(e.get("amount")))])

    out = pdf.output()
    return bytes(out)


# =============================================================
#  Excel
# =============================================================
def build_xlsx(state):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    c = compute(state)
    event = state.get("event", {}) or {}
    MONEY = '#,##0'
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="17352C")
    bold = Font(bold=True)

    wb = Workbook()

    # ---- Resumen ----
    ws = wb.active
    ws.title = "Resumen"
    rows = [
        ("Evento", event.get("name", "")),
        ("Lugar", event.get("place", "")),
        ("Fecha", event.get("date", "")),
        ("Valor poldina", c["pv"]),
        ("", ""),
        ("Total poldinas", c["totalPoldinas"]),
        ("Bolsa de multas", c["pool"]),
        ("Consumo comun", c["sharedCost"]),
        ("Exceso repartido", c["excess"]),
        ("Multas sin usar", c["surplus"]),
        ("Extras individuales", c["extrasTotal"]),
        ("TOTAL a recaudar", c["grand"]),
    ]
    for k, v in rows:
        ws.append([k, v])
    for cell in ws["B"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY
    ws["A12"].font = bold
    ws["B12"].font = bold
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22

    # ---- Cada quien paga ----
    ws2 = wb.create_sheet("Cada quien paga")
    headers = ["Persona", "Poldinas", "Multas", "Exceso", "Extras", "Total"]
    ws2.append(headers)
    for i, cell in enumerate(ws2[1], 1):
        cell.font = head_font
        cell.fill = head_fill
    for r in c["rows"]:
        ws2.append([r["name"], r["poldinas"], r["fine"], r["excessShare"], r["extras"], r["total"]])
    ws2.append(["TOTAL", "", sum(r["fine"] for r in c["rows"]), c["excess"], c["extrasTotal"], c["grand"]])
    for col in ("C", "D", "E", "F"):
        for cell in ws2[col]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    for cell in ws2[ws2.max_row]:
        cell.font = bold
    ws2.column_dimensions["A"].width = 22
    for col in ("B", "C", "D", "E", "F"):
        ws2.column_dimensions[col].width = 13

    # ---- Consumo comun ----
    ws3 = wb.create_sheet("Consumo comun")
    ws3.append(["Item", "Cantidad", "Precio unit", "Subtotal"])
    for cell in ws3[1]:
        cell.font = head_font
        cell.fill = head_fill
    for i in (state.get("shared", []) or []):
        q, p = num(i.get("qty")), num(i.get("unitPrice"))
        ws3.append([i.get("desc", ""), inum(i.get("qty")), p, q * p])
    for col in ("C", "D"):
        for cell in ws3[col]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    ws3.column_dimensions["A"].width = 26
    for col in ("B", "C", "D"):
        ws3.column_dimensions[col].width = 13

    # ---- Extras ----
    ws4 = wb.create_sheet("Extras")
    ws4.append(["Persona", "Concepto", "Monto"])
    for cell in ws4[1]:
        cell.font = head_font
        cell.fill = head_fill
    name_of = {p.get("id"): (p.get("name") or "(sin nombre)") for p in state.get("people", [])}
    for e in (state.get("extras", []) or []):
        ws4.append([name_of.get(e.get("personId"), "(sin asignar)"), e.get("desc", ""), num(e.get("amount"))])
    for cell in ws4["C"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 26
    ws4.column_dimensions["C"].width = 13

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def filename(state, ext):
    base = (state.get("event", {}) or {}).get("name") or "poldinas"
    safe = "".join(ch if (ch.isalnum() or ch in "-_") else "_" for ch in base).strip("_").lower() or "poldinas"
    return f"{safe}.{ext}"


# =============================================================
#  Rutas
# =============================================================
@app.route("/")
def home():
    return Response(PAGE, mimetype="text/html")


@app.route("/api/calc", methods=["POST"])
def api_calc():
    state = request.get_json(force=True, silent=True) or {}
    return jsonify(compute(state))


@app.route("/api/export/pdf", methods=["POST"])
def api_pdf():
    state = request.get_json(force=True, silent=True) or {}
    try:
        data = build_pdf(state)
    except Exception as e:
        return jsonify({"error": f"No se pudo generar el PDF: {e}"}), 500
    return send_file(BytesIO(data), mimetype="application/pdf",
                     as_attachment=True, download_name=filename(state, "pdf"))


@app.route("/api/export/xlsx", methods=["POST"])
def api_xlsx():
    state = request.get_json(force=True, silent=True) or {}
    try:
        data = build_xlsx(state)
    except Exception as e:
        return jsonify({"error": f"No se pudo generar el Excel: {e}"}), 500
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename(state, "xlsx"))


# =============================================================
#  Interfaz (HTML + CSS + JS)
# =============================================================
PAGE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Poldinas - repartidor de la cuenta</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500;600;700&display=swap" rel="stylesheet">
<style>
  :root{
    --paper:#E7ECE2; --card:#FBFBF5; --card-2:#F2F4EC;
    --ink:#17352C; --ink-soft:#54655C; --line:#D4DACC;
    --brass:#A87A1E; --brass-deep:#815d13; --brass-soft:#F0E6CE;
    --brick:#C24E33; --brick-soft:#F4E1D9;
    --pine:#2A6A54; --pine-soft:#DBEBE1;
    --radius:14px; --shadow:0 1px 2px rgba(23,53,44,.06), 0 8px 24px rgba(23,53,44,.06);
  }
  *{box-sizing:border-box}
  html,body{margin:0}
  body{background:var(--paper); color:var(--ink); font-family:"Inter",system-ui,sans-serif; font-size:15px; line-height:1.5; -webkit-font-smoothing:antialiased;}
  .wrap{max-width:1120px; margin:0 auto; padding:26px 20px 80px;}
  header.top{display:flex; align-items:flex-end; gap:16px; flex-wrap:wrap; margin-bottom:22px;}
  .brand{display:flex; align-items:center; gap:14px;}
  .coin{width:52px; height:52px; border-radius:50%; flex:none;
    background:radial-gradient(circle at 34% 30%, #E7C878, var(--brass) 62%, var(--brass-deep));
    box-shadow:inset 0 2px 4px rgba(255,255,255,.5), inset 0 -3px 6px rgba(0,0,0,.25), var(--shadow);
    display:grid; place-items:center; color:#4a3608; font-family:"Fraunces",serif; font-weight:700; font-size:24px; border:2px solid #caa14e;}
  h1{font-family:"Fraunces",serif; font-weight:600; font-size:30px; line-height:1; margin:0; letter-spacing:-.01em;}
  .tag{color:var(--ink-soft); font-size:13px; margin-top:5px;}
  .top-actions{margin-left:auto; display:flex; gap:8px; flex-wrap:wrap;}
  .grid{display:grid; grid-template-columns:1fr 400px; gap:20px; align-items:start;}
  @media(max-width:900px){ .grid{grid-template-columns:1fr;} .side{position:static !important;} }
  .side{position:sticky; top:20px;}
  .card{background:var(--card); border:1px solid var(--line); border-radius:var(--radius); box-shadow:var(--shadow); margin-bottom:18px; overflow:hidden;}
  .card-h{display:flex; align-items:center; gap:10px; padding:15px 18px 13px; border-bottom:1px solid var(--line);}
  .card-h .n{font-family:"JetBrains Mono",monospace; font-size:12px; color:var(--brass-deep); background:var(--brass-soft); border-radius:6px; padding:2px 7px; font-weight:700;}
  .card-h h2{font-family:"Fraunces",serif; font-weight:600; font-size:18px; margin:0; flex:1;}
  .card-h .sub{color:var(--ink-soft); font-size:12.5px;}
  .card-b{padding:16px 18px;}
  label{display:block; font-size:12px; font-weight:600; color:var(--ink-soft); margin:0 0 5px;}
  input,select{width:100%; font-family:inherit; font-size:14px; color:var(--ink); background:#fff; border:1px solid var(--line); border-radius:9px; padding:9px 11px; outline:none;}
  input:focus,select:focus{border-color:var(--brass); box-shadow:0 0 0 3px var(--brass-soft);}
  input[type=number]{font-family:"JetBrains Mono",monospace;}
  .field-row{display:grid; gap:10px;}
  .cols-3{grid-template-columns:1.6fr 1fr auto;}
  @media(max-width:560px){ .cols-3{grid-template-columns:1fr;} }
  button{font-family:inherit; cursor:pointer; border:none; border-radius:9px; font-weight:600; font-size:13.5px; padding:9px 14px; transition:filter .12s, background .12s;}
  button:hover{filter:brightness(.97)}
  .btn-primary{background:var(--ink); color:#fff;}
  .btn-brass{background:var(--brass); color:#fff;}
  .btn-ghost{background:transparent; color:var(--ink); border:1px solid var(--line);}
  .btn-ghost:hover{background:var(--card-2)}
  .btn-sm{padding:7px 11px; font-size:12.5px;}
  .btn-x{background:transparent; color:var(--ink-soft); border:1px solid var(--line); padding:8px 10px; border-radius:8px; line-height:1;}
  .btn-x:hover{background:var(--brick-soft); color:var(--brick); border-color:var(--brick-soft);}
  .row{display:grid; gap:9px; align-items:center; padding:10px 0; border-top:1px dashed var(--line);}
  .row:first-child{border-top:none;}
  .row-people{grid-template-columns:1fr 130px auto;}
  .row-shared{grid-template-columns:1fr 70px 1fr 90px auto;}
  .row-extra{grid-template-columns:1fr 1fr 110px auto;}
  @media(max-width:560px){ .row-people,.row-shared,.row-extra{grid-template-columns:1fr 1fr;} .row .span{grid-column:1/-1;} }
  .row .lineamt{font-family:"JetBrains Mono",monospace; font-size:13px; text-align:right; color:var(--ink-soft); align-self:center; white-space:nowrap;}
  .empty{color:var(--ink-soft); font-size:13.5px; text-align:center; padding:16px 8px; font-style:italic;}
  .meter-wrap{margin:6px 0 14px;}
  .meter-title{font-size:11px; text-transform:uppercase; letter-spacing:.08em; color:var(--ink-soft); margin-bottom:7px; font-weight:600;}
  .meter{height:26px; border-radius:8px; background:var(--card-2); border:1px solid var(--line); display:flex; overflow:hidden;}
  .seg{height:100%; transition:width .3s ease;}
  .seg.covered{background:repeating-linear-gradient(45deg,var(--pine),var(--pine) 10px,#2f7a61 10px,#2f7a61 20px);}
  .seg.excess{background:repeating-linear-gradient(45deg,var(--brick),var(--brick) 8px,#d15b40 8px,#d15b40 16px);}
  .seg.surplus{background:repeating-linear-gradient(45deg,var(--brass-soft),var(--brass-soft) 8px,#e6d6a8 8px,#e6d6a8 16px);}
  .legend{display:flex; flex-wrap:wrap; gap:12px; margin-top:9px; font-size:11.5px; color:var(--ink-soft);}
  .legend span{display:inline-flex; align-items:center; gap:6px;}
  .dot{width:11px;height:11px;border-radius:3px; flex:none;}
  .dot.c{background:var(--pine)} .dot.e{background:var(--brick)} .dot.s{background:#e6d6a8}
  .stat-line{display:flex; justify-content:space-between; align-items:baseline; padding:7px 0; border-top:1px solid var(--line); font-size:13.5px;}
  .stat-line:first-of-type{border-top:none;}
  .stat-line .k{color:var(--ink-soft);}
  .stat-line .v{font-family:"JetBrains Mono",monospace; font-weight:600;}
  .stat-line.big{padding-top:12px;}
  .stat-line.big .k{font-family:"Fraunces",serif; font-size:16px; color:var(--ink);}
  .stat-line.big .v{font-size:19px; color:var(--brass-deep);}
  .note{font-size:12.5px; border-radius:10px; padding:10px 12px; margin-top:12px; line-height:1.45;}
  .note.warn{background:var(--brick-soft); color:#8a3620;}
  .note.info{background:var(--pine-soft); color:#1f4d3c;}
  table.tbl{width:100%; border-collapse:collapse; font-size:13px;}
  table.tbl th{text-align:right; font-size:11px; text-transform:uppercase; letter-spacing:.05em; color:var(--ink-soft); font-weight:600; padding:8px 8px; border-bottom:1px solid var(--line);}
  table.tbl th:first-child{text-align:left;}
  table.tbl td{padding:9px 8px; border-bottom:1px solid var(--line); text-align:right; font-family:"JetBrains Mono",monospace;}
  table.tbl td:first-child{text-align:left; font-family:"Inter",sans-serif; font-weight:500;}
  table.tbl tr.total-row td{border-top:2px solid var(--ink); border-bottom:none; font-weight:700; padding-top:11px;}
  table.tbl tr.total-row td:first-child{font-family:"Fraunces",serif;}
  .pill{display:inline-block; font-size:10.5px; padding:1px 7px; border-radius:20px; font-weight:600; margin-left:6px; vertical-align:middle;}
  .pill.fined{background:var(--brass-soft); color:var(--brass-deep);}
  .pill.free{background:var(--pine-soft); color:var(--pine);}
  .grand{font-family:"Fraunces",serif; font-weight:700; color:var(--brass-deep);}
  .hist-item{display:flex; align-items:center; gap:10px; padding:9px 0; border-top:1px dashed var(--line); font-size:13px;}
  .hist-item:first-child{border-top:none;}
  .hist-item .hname{flex:1; font-weight:500;}
  .hist-item .hdate{color:var(--ink-soft); font-size:11.5px; font-family:"JetBrains Mono",monospace;}
  .toast{position:fixed; left:50%; bottom:26px; transform:translateX(-50%) translateY(20px); background:var(--ink); color:#fff; padding:11px 18px; border-radius:10px; font-size:13.5px; box-shadow:var(--shadow); opacity:0; transition:.25s; pointer-events:none; z-index:50;}
  .toast.show{opacity:1; transform:translateX(-50%) translateY(0);}
  .save-flag{font-size:11.5px; color:var(--ink-soft); margin-left:auto; display:flex; align-items:center; gap:6px;}
  .save-flag .d{width:7px;height:7px;border-radius:50%;background:var(--pine);}
</style>
</head>
<body>
<div class="wrap">
  <header class="top">
    <div class="brand">
      <div class="coin">P</div>
      <div><h1>Poldinas</h1><div class="tag">Reparte la cuenta: multas, exceso y extras - cada quien lo suyo.</div></div>
    </div>
    <div class="top-actions">
      <button class="btn-ghost btn-sm" id="btnNew">Nuevo evento</button>
      <button class="btn-primary btn-sm" id="btnSave">Guardar evento</button>
    </div>
  </header>

  <div class="grid">
    <div class="main">
      <section class="card">
        <div class="card-h"><span class="n">01</span><h2>Evento</h2>
          <div class="save-flag"><span class="d"></span> <span id="saveState">Guardado</span></div></div>
        <div class="card-b">
          <div class="field-row cols-3">
            <div><label>Nombre / motivo</label><input id="evName" placeholder="Ej: Poldinas de julio"></div>
            <div><label>Lugar</label><input id="evPlace" placeholder="Ej: Pizzeria del centro"></div>
            <div style="width:130px"><label>Valor poldina</label><input id="evPV" type="number" min="0" step="100" value="1000"></div>
          </div>
          <div class="field-row" style="margin-top:10px; grid-template-columns:1fr; max-width:220px">
            <div><label>Fecha</label><input id="evDate" type="date"></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">02</span><h2>Personas</h2><span class="sub">Pon 0 poldinas a quien no tiene multa</span></div>
        <div class="card-b">
          <div id="peopleList"></div>
          <div class="field-row cols-3" style="margin-top:14px">
            <div><label>Nombre</label><input id="pName" placeholder="Nombre de la persona"></div>
            <div><label>Poldinas</label><input id="pPold" type="number" min="0" step="1" value="0"></div>
            <div style="display:flex; align-items:flex-end"><button class="btn-brass" id="btnAddPerson" style="width:100%">Agregar</button></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">03</span><h2>Consumo comun (factura)</h2><span class="sub">Lo que cubren las multas</span></div>
        <div class="card-b">
          <div id="sharedList"></div>
          <div class="field-row" style="margin-top:14px; grid-template-columns:1.4fr 1fr 1fr .8fr; gap:10px">
            <div><label>Item</label><input id="sDesc" placeholder="Ej: Pizza mediana"></div>
            <div><label>Cant.</label><input id="sQty" type="number" min="0" step="1" value="1"></div>
            <div><label>Precio unit.</label><input id="sPrice" type="number" min="0" step="100" placeholder="1000"></div>
            <div style="display:flex; align-items:flex-end"><button class="btn-brass" id="btnAddShared" style="width:100%">Agregar</button></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">04</span><h2>Consumos individuales</h2><span class="sub">Bebidas y extras - los paga cada quien</span></div>
        <div class="card-b">
          <div id="extraList"></div>
          <div class="field-row" style="margin-top:14px; grid-template-columns:1.2fr 1.4fr 1fr auto; gap:10px">
            <div><label>Persona</label><select id="xPerson"></select></div>
            <div><label>Concepto</label><input id="xDesc" placeholder="Ej: Gaseosa, cerveza"></div>
            <div><label>Monto</label><input id="xAmt" type="number" min="0" step="100" placeholder="0"></div>
            <div style="display:flex; align-items:flex-end"><button class="btn-brass" id="btnAddExtra" style="width:100%">Agregar</button></div>
          </div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><span class="n">05</span><h2>Eventos guardados</h2></div>
        <div class="card-b"><div id="histList"></div></div>
      </section>
    </div>

    <div class="side">
      <section class="card">
        <div class="card-h"><h2 style="font-size:20px">Resumen</h2></div>
        <div class="card-b" style="padding:18px">
          <div class="meter-wrap">
            <div class="meter-title">Bolsa de multas vs. consumo comun</div>
            <div class="meter" id="meter"></div>
            <div class="legend">
              <span><i class="dot c"></i>Cubierto por multas</span>
              <span><i class="dot e"></i>Exceso (sin multa)</span>
              <span><i class="dot s"></i>Multas sin usar</span>
            </div>
          </div>
          <div class="stat-line"><span class="k">Bolsa de multas <span id="poldCount" class="mono"></span></span><span class="v" id="vPool">$0</span></div>
          <div class="stat-line"><span class="k">Consumo comun</span><span class="v" id="vShared">$0</span></div>
          <div class="stat-line"><span class="k">Exceso a repartir</span><span class="v" id="vExcess" style="color:var(--brick)">$0</span></div>
          <div class="stat-line"><span class="k">Extras individuales</span><span class="v" id="vExtras">$0</span></div>
          <div class="stat-line big"><span class="k">Total a recaudar</span><span class="v" id="vGrand">$0</span></div>
          <div id="noteBox"></div>
        </div>
      </section>

      <section class="card">
        <div class="card-h"><h2 style="font-size:18px">Cada quien paga</h2></div>
        <div class="card-b" style="padding:10px 12px 14px">
          <div id="breakdown"></div>
          <div style="display:flex; gap:8px; margin-top:16px;">
            <button class="btn-primary" id="btnPDF" style="flex:1">Exportar PDF</button>
            <button class="btn-brass" id="btnXLS" style="flex:1">Exportar Excel</button>
          </div>
        </div>
      </section>
    </div>
  </div>
</div>
<div class="toast" id="toast"></div>

<script>
const uid = () => Date.now().toString(36) + Math.random().toString(36).slice(2,7);
const CUR = new Intl.NumberFormat('es-CO',{style:'currency',currency:'COP',maximumFractionDigits:0});
const money = n => CUR.format(Math.round(n||0));

const LS_CUR='poldinas:current', LS_EVENTS='poldinas:events';
let state = loadCurrent() || blank();

function blank(){ return {id:uid(), event:{name:'',place:'',date:'',poldinaValue:1000}, people:[], shared:[], extras:[]}; }
function loadCurrent(){ try{ const r=localStorage.getItem(LS_CUR); return r?JSON.parse(r):null; }catch(e){ return null; } }
function loadEvents(){ try{ const r=localStorage.getItem(LS_EVENTS); return r?JSON.parse(r):[]; }catch(e){ return []; } }

let saveT=null;
function save(){ const s=document.getElementById('saveState'); if(s) s.textContent='Guardando...';
  clearTimeout(saveT); saveT=setTimeout(()=>{ try{ localStorage.setItem(LS_CUR, JSON.stringify(state)); if(s) s.textContent='Guardado'; }catch(e){ if(s) s.textContent='No se pudo guardar'; } }, 300); }

/* ---------- calculo: pide a Python, con respaldo local ---------- */
let calcT=null;
function scheduleCalc(){ clearTimeout(calcT); calcT=setTimeout(doCalc, 250); }
async function doCalc(){
  try{
    const res = await fetch('/api/calc',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(state)});
    if(!res.ok) throw 0;
    renderSummary(await res.json());
  }catch(e){ renderSummary(jsCalc()); }
}
function jsCalc(){
  const pv=+state.event.poldinaValue||1000;
  const totalPold=state.people.reduce((s,p)=>s+(+p.poldinas||0),0);
  const pool=totalPold*pv;
  const sharedCost=state.shared.reduce((s,i)=>s+(+i.qty||0)*(+i.unitPrice||0),0);
  const excess=Math.max(0,sharedCost-pool), surplus=Math.max(0,pool-sharedCost), covered=Math.min(sharedCost,pool);
  const nonF=state.people.filter(p=>(+p.poldinas||0)===0);
  let base=nonF, fallback=false; if(excess>0&&!nonF.length){ base=state.people; fallback=true; }
  const per=base.length?excess/base.length:0; const ids=new Set(base.map(p=>p.id));
  const rows=state.people.map(p=>{ const fine=(+p.poldinas||0)*pv;
    const ex=state.extras.filter(e=>e.personId===p.id).reduce((s,e)=>s+(+e.amount||0),0);
    const sh=(excess>0&&ids.has(p.id))?per:0;
    return {id:p.id,name:p.name||'(sin nombre)',poldinas:+p.poldinas||0,fine,excessShare:sh,extras:ex,total:fine+ex+sh,fined:(+p.poldinas||0)>0}; });
  const extrasTotal=state.extras.reduce((s,e)=>s+(+e.amount||0),0);
  return {totalPoldinas:totalPold,pool,sharedCost,covered,excess,surplus,extrasTotal,grand:rows.reduce((s,r)=>s+r.total,0),excessPer:per,nonFinedCount:nonF.length,fallback,rows};
}

/* ---------- helpers DOM ---------- */
function el(tag, attrs={}, kids=[]){ const n=document.createElement(tag);
  for(const k in attrs){ if(k==='class')n.className=attrs[k]; else if(k==='html')n.innerHTML=attrs[k];
    else if(k.startsWith('on'))n.addEventListener(k.slice(2),attrs[k]); else n.setAttribute(k,attrs[k]); }
  (Array.isArray(kids)?kids:[kids]).forEach(c=>{ if(c!=null)n.append(c.nodeType?c:document.createTextNode(c)); });
  return n; }

/* ---------- listas ---------- */
function renderPeople(){ const box=document.getElementById('peopleList'); box.innerHTML='';
  if(!state.people.length){ box.append(el('div',{class:'empty'},'Aun no hay personas.')); return; }
  state.people.forEach(p=>{
    const name=el('input',{value:p.name||'',placeholder:'Nombre',class:'span',oninput:e=>{p.name=e.target.value; refreshExtraSelect(); scheduleCalc(); save();}});
    const pold=el('input',{type:'number',min:'0',step:'1',value:p.poldinas??0,oninput:e=>{p.poldinas=+e.target.value||0; scheduleCalc(); save();}});
    const del=el('button',{class:'btn-x',title:'Quitar',onclick:()=>{ state.people=state.people.filter(x=>x.id!==p.id); state.extras=state.extras.filter(x=>x.personId!==p.id); renderPeople(); renderExtras(); refreshExtraSelect(); scheduleCalc(); save(); }},'X');
    box.append(el('div',{class:'row row-people'},[name,pold,del]));
  });
}
function renderShared(){ const box=document.getElementById('sharedList'); box.innerHTML='';
  if(!state.shared.length){ box.append(el('div',{class:'empty'},'Sin items de consumo comun.')); return; }
  state.shared.forEach(it=>{
    const amt=el('div',{class:'lineamt'}, money((+it.qty||0)*(+it.unitPrice||0)));
    const upd=()=>{ amt.textContent=money((+it.qty||0)*(+it.unitPrice||0)); };
    const desc=el('input',{value:it.desc||'',placeholder:'Item',class:'span',oninput:e=>{it.desc=e.target.value; save();}});
    const qty=el('input',{type:'number',min:'0',step:'1',value:it.qty??1,oninput:e=>{it.qty=+e.target.value||0; upd(); scheduleCalc(); save();}});
    const price=el('input',{type:'number',min:'0',step:'100',value:it.unitPrice??0,oninput:e=>{it.unitPrice=+e.target.value||0; upd(); scheduleCalc(); save();}});
    const del=el('button',{class:'btn-x',title:'Quitar',onclick:()=>{ state.shared=state.shared.filter(x=>x.id!==it.id); renderShared(); scheduleCalc(); save(); }},'X');
    box.append(el('div',{class:'row row-shared'},[desc,qty,price,amt,del]));
  });
}
function renderExtras(){ const box=document.getElementById('extraList'); box.innerHTML='';
  if(!state.extras.length){ box.append(el('div',{class:'empty'},'Sin consumos individuales.')); return; }
  state.extras.forEach(x=>{
    const sel=el('select',{onchange:e=>{x.personId=e.target.value; scheduleCalc(); save();}}); fillOptions(sel,x.personId);
    const desc=el('input',{value:x.desc||'',placeholder:'Concepto',oninput:e=>{x.desc=e.target.value; save();}});
    const amt=el('input',{type:'number',min:'0',step:'100',value:x.amount??0,oninput:e=>{x.amount=+e.target.value||0; scheduleCalc(); save();}});
    const del=el('button',{class:'btn-x',title:'Quitar',onclick:()=>{ state.extras=state.extras.filter(v=>v.id!==x.id); renderExtras(); scheduleCalc(); save(); }},'X');
    box.append(el('div',{class:'row row-extra'},[sel,desc,amt,del]));
  });
}
function fillOptions(sel,selected){ sel.innerHTML='';
  if(!state.people.length){ sel.append(el('option',{value:''},'- agrega personas -')); return; }
  state.people.forEach(p=>{ const o=el('option',{value:p.id}, p.name||'(sin nombre)'); if(p.id===selected)o.selected=true; sel.append(o); }); }
function refreshExtraSelect(){ const add=document.getElementById('xPerson'); fillOptions(add,add.value); renderExtras(); }

/* ---------- resumen ---------- */
function renderSummary(c){ renderMeter(c); renderStats(c); renderBreakdown(c); }
function renderMeter(c){ const m=document.getElementById('meter'); m.innerHTML='';
  const scale=Math.max(c.pool,c.sharedCost,1), w=v=>Math.max(0,(v/scale)*100);
  m.append(el('div',{class:'seg covered',style:`width:${w(c.covered)}%`,title:'Cubierto: '+money(c.covered)}));
  if(c.excess>0) m.append(el('div',{class:'seg excess',style:`width:${w(c.excess)}%`,title:'Exceso: '+money(c.excess)}));
  if(c.surplus>0) m.append(el('div',{class:'seg surplus',style:`width:${w(c.surplus)}%`,title:'Sin usar: '+money(c.surplus)}));
}
function renderStats(c){
  document.getElementById('poldCount').textContent='('+c.totalPoldinas+' pold.)';
  document.getElementById('vPool').textContent=money(c.pool);
  document.getElementById('vShared').textContent=money(c.sharedCost);
  document.getElementById('vExcess').textContent=money(c.excess);
  document.getElementById('vExtras').textContent=money(c.extrasTotal);
  document.getElementById('vGrand').textContent=money(c.grand);
  const nb=document.getElementById('noteBox'); nb.innerHTML='';
  if(c.excess>0 && !c.fallback) nb.append(el('div',{class:'note warn'},`El consumo se paso por ${money(c.excess)}. Se reparte entre ${c.nonFinedCount} persona(s) sin multa -> ${money(c.excessPer)} c/u.`));
  if(c.fallback) nb.append(el('div',{class:'note warn'},`Hay un exceso de ${money(c.excess)} pero nadie esta sin multa. Se repartio entre todas las personas (${money(c.excessPer)} c/u).`));
  if(c.surplus>0) nb.append(el('div',{class:'note info'},`Las multas superan el consumo comun por ${money(c.surplus)} (bolsa sin usar).`));
}
function renderBreakdown(c){ const box=document.getElementById('breakdown'); box.innerHTML='';
  if(!c.rows.length){ box.append(el('div',{class:'empty'},'Agrega personas para ver el reparto.')); return; }
  const tbl=el('table',{class:'tbl'});
  tbl.append(el('thead',{},el('tr',{},[el('th',{},'Persona'),el('th',{},'Multas'),el('th',{},'Exceso'),el('th',{},'Extras'),el('th',{},'Total')])));
  const tb=el('tbody');
  c.rows.forEach(r=>{
    const pill = r.fined ? el('span',{class:'pill fined'}, r.poldinas+'p') : el('span',{class:'pill free'},'sin multa');
    tb.append(el('tr',{},[ el('td',{},[document.createTextNode(r.name),pill]),
      el('td',{},money(r.fine)), el('td',{}, r.excessShare?money(r.excessShare):'-'),
      el('td',{}, r.extras?money(r.extras):'-'), el('td',{class:'grand'},money(r.total)) ]));
  });
  tb.append(el('tr',{class:'total-row'},[ el('td',{},'Total'),
    el('td',{},money(c.rows.reduce((s,r)=>s+r.fine,0))), el('td',{},money(c.excess)),
    el('td',{},money(c.extrasTotal)), el('td',{class:'grand'},money(c.grand)) ]));
  tbl.append(tb); box.append(tbl);
}

/* ---------- historial ---------- */
function renderHistory(){ const box=document.getElementById('histList'); box.innerHTML='';
  const evs=loadEvents().sort((a,b)=>(b.savedAt||'').localeCompare(a.savedAt||''));
  if(!evs.length){ box.append(el('div',{class:'empty'},'Todavia no has guardado eventos.')); return; }
  evs.forEach(ev=>{
    const d=ev.savedAt?new Date(ev.savedAt).toLocaleDateString('es-CO',{day:'2-digit',month:'short',year:'2-digit'}):'';
    const load=el('button',{class:'btn-ghost btn-sm',onclick:()=>{ state=JSON.parse(JSON.stringify(ev.state)); syncInputs(); renderAll(); save(); toast('Evento cargado'); }},'Cargar');
    const del=el('button',{class:'btn-x',title:'Eliminar',onclick:()=>{ const list=loadEvents().filter(e=>e.id!==ev.id); localStorage.setItem(LS_EVENTS,JSON.stringify(list)); renderHistory(); toast('Evento eliminado'); }},'X');
    box.append(el('div',{class:'hist-item'},[ el('span',{class:'hname'}, ev.name||'Evento'), el('span',{class:'hdate'}, d), load, del ]));
  });
}

/* ---------- inputs evento ---------- */
function syncInputs(){
  document.getElementById('evName').value=state.event.name||'';
  document.getElementById('evPlace').value=state.event.place||'';
  document.getElementById('evDate').value=state.event.date||'';
  document.getElementById('evPV').value=state.event.poldinaValue??1000;
}
function renderAll(){ renderPeople(); renderShared(); refreshExtraSelect(); scheduleCalc(); }

/* ---------- exportar ---------- */
async function download(url){
  try{
    const res=await fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(state)});
    if(!res.ok) throw 0;
    const blob=await res.blob(); const dl=URL.createObjectURL(blob);
    const cd=res.headers.get('Content-Disposition')||''; const m=cd.match(/filename="?([^"]+)"?/);
    const a=document.createElement('a'); a.href=dl; a.download=m?m[1]:'poldinas'; document.body.append(a); a.click(); a.remove(); URL.revokeObjectURL(dl);
  }catch(e){ toast('No se pudo exportar (revisa la conexion)'); }
}

/* ---------- toast ---------- */
let tT=null; function toast(msg){ const t=document.getElementById('toast'); t.textContent=msg; t.classList.add('show'); clearTimeout(tT); tT=setTimeout(()=>t.classList.remove('show'),2200); }

/* ---------- bind ---------- */
function bind(){ const on=(id,ev,fn)=>document.getElementById(id).addEventListener(ev,fn);
  on('evName','input',e=>{state.event.name=e.target.value; save();});
  on('evPlace','input',e=>{state.event.place=e.target.value; save();});
  on('evDate','input',e=>{state.event.date=e.target.value; save();});
  on('evPV','input',e=>{state.event.poldinaValue=+e.target.value||0; scheduleCalc(); save();});
  on('btnAddPerson','click',()=>{ const n=document.getElementById('pName').value.trim(); if(!n){toast('Escribe un nombre');return;}
    state.people.push({id:uid(),name:n,poldinas:+document.getElementById('pPold').value||0});
    document.getElementById('pName').value=''; document.getElementById('pPold').value=0;
    renderPeople(); refreshExtraSelect(); scheduleCalc(); save(); document.getElementById('pName').focus(); });
  on('btnAddShared','click',()=>{ const d=document.getElementById('sDesc').value.trim(); if(!d){toast('Escribe el item');return;}
    state.shared.push({id:uid(),desc:d,qty:+document.getElementById('sQty').value||0,unitPrice:+document.getElementById('sPrice').value||0});
    document.getElementById('sDesc').value=''; document.getElementById('sQty').value=1; document.getElementById('sPrice').value='';
    renderShared(); scheduleCalc(); save(); document.getElementById('sDesc').focus(); });
  on('btnAddExtra','click',()=>{ if(!state.people.length){toast('Primero agrega personas');return;}
    const d=document.getElementById('xDesc').value.trim(); if(!d){toast('Escribe el concepto');return;}
    state.extras.push({id:uid(),personId:document.getElementById('xPerson').value,desc:d,amount:+document.getElementById('xAmt').value||0});
    document.getElementById('xDesc').value=''; document.getElementById('xAmt').value='';
    renderExtras(); scheduleCalc(); save(); document.getElementById('xDesc').focus(); });
  on('btnSave','click',()=>{ const list=loadEvents(); const snap={id:state.id,name:state.event.name||'Evento sin nombre',savedAt:new Date().toISOString(),state:JSON.parse(JSON.stringify(state))};
    const i=list.findIndex(e=>e.id===state.id); if(i>=0)list[i]=snap; else list.push(snap);
    localStorage.setItem(LS_EVENTS,JSON.stringify(list)); renderHistory(); toast('Evento guardado'); });
  on('btnNew','click',()=>{ state=blank(); syncInputs(); renderAll(); save(); toast('Evento nuevo'); });
  on('btnPDF','click',()=>download('/api/export/pdf'));
  on('btnXLS','click',()=>download('/api/export/xlsx'));
}

/* ---------- init ---------- */
bind(); syncInputs(); renderAll(); renderHistory();
</script>
</body>
</html>"""

if __name__ == "__main__":
    app.run(debug=True, port=5000)
